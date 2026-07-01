"""
launch_SQL_query_V2.py
----------------------
Version industrialisee du script d'execution des requetes SQL Oracle
(MDG/DWH) et PostgreSQL (IEHE).

Idee directrice : un utilisateur non technique doit pouvoir ajouter une
requete, une categorie ou un environnement sans toucher au code Python.

Prerequis Python :
    pip install oracledb pandas openpyxl
    pip install "psycopg[binary]"            (v3, prefere)
        ou en fallback :
    pip install psycopg2-binary              (v2)
psycopg n'est requis que si des requetes IEHE sont selectionnees.

Configuration externalisee :
  * `input_data/SQL/environments.ini` liste les couples <ENV>:<base> et
    leur configuration de connexion :
        Oracle    (mdg, dwh)  -> host, port, sid
        Postgres  (iehe)      -> host, port, dbname
    Ajouter une section => nouvel environnement detecte automatiquement.
  * `input_data/SQL/<CATEGORIE>/*.sql` regroupe les requetes par theme.
    Le sous-dossier devient automatiquement la categorie proposee au
    menu. Ajouter un dossier => nouvelle categorie.
  * Identifiants Oracle saisis depuis le .bat parent et transmis via
    ORA_USER / ORA_PASSWORD.
  * Identifiants IEHE optionnels via IEHE_USER / IEHE_PASSWORD ; en
    leur absence, on reutilise les identifiants Oracle.

METADONNEES SQL
---------------
Chaque `.sql` porte ses metadonnees en en-tete sous forme de commentaires :

    -- db: mdg                            (ou "dwh", "iehe", ...)
    -- post_process: sum_total_doublons   (optionnel)
    -- param.<NOM>: <prompt> [<defaut>]   (optionnel, repetable)

Le bloc de metadonnees s'arrete a la premiere ligne non vide qui n'est
pas un de ces commentaires.

PARAMETRES UTILISATEUR (`-- param.X:`)
--------------------------------------
Format :
    -- param.NOM_PARAM: Texte du prompt affiche [defaut_optionnel]

Regles :
  * NOM_PARAM : majuscules ASCII + chiffres + `_`, commence par une lettre.
  * `[defaut]` est facultatif. Tokens speciaux :
        [TODAY]      -> date du jour    au format YYYYMMDD
        [YESTERDAY]  -> date d'hier     au format YYYYMMDD
        [autre]      -> valeur litterale
  * Le SQL utilise `{NOM_PARAM}` (regex ^\\{[A-Z][A-Z0-9_]*\\}$) ;
    la substitution remplace cette syntaxe par la valeur saisie.
  * Les apostrophes des valeurs sont doublees avant substitution pour
    eviter de casser un litteral SQL `'{NOM}'`. Le quoting (apostrophes
    autour du placeholder) reste a la charge du SQL.
  * Validation au chargement : tout `{NOM}` utilise dans le SQL sans
    declaration `-- param.NOM:` provoque une erreur fatale AVANT toute
    connexion DB. Exception : `{DATE_SUIVI}` est tolere par compat
    (substitue automatiquement par la date de frequence saisie).

CONVENTION DE NOMMAGE — frequence et inclusion
----------------------------------------------
Le PREFIXE du nom de fichier indique la frequence :
  * `J_<nom>.sql` -> requete QUOTIDIENNE (Jour),  auto-executee
  * `S_<nom>.sql` -> requete HEBDOMADAIRE (Semaine), auto-executee
  * tout autre nom (sans prefixe `J_` ou `S_`) -> IGNORE

Le sous-dossier `DDL/` est reserve aux scripts de creation/maintenance
et n'est pas inclus dans la liste des requetes a executer.

Workflow runtime :
  1) Menu multi-selection des categories detectees (ex. "1,3" ou "A").
  2) Choix de l'environnement (parmi ceux declares dans environments.ini).
  3) Saisie de la date de suivi par frequence presente (DD/MM/YYYY).
  3bis) Saisie des parametres user declares dans les .sql selectionnes
        (un prompt unique par parametre, partage entre toutes les requetes).
  4) Execution parallele -> un xlsx par categorie sous
     output/SQL/<date>/<CATEGORIE>_<HHMMSS>_Resultats_SQL.xlsx.
"""

from __future__ import annotations

import configparser
import getpass
import os
import re
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import oracledb

# T1 — helpers Prestations_Par_Offre (post-processor + builder Excel pivot)
import prestations_par_offre_lib as ppo

# psycopg est requis pour la base IEHE (PostgreSQL). Import conditionnel
# pour ne pas bloquer le script si seules des requetes Oracle (mdg/dwh)
# sont a executer. La validation effective (presence + capacite a se
# connecter) est faite au moment ou une requete IEHE est selectionnee.
try:
    import psycopg as _psycopg  # v3 (paquet 'psycopg')
    _PSYCOPG_VERSION = 3
except ImportError:
    try:
        import psycopg2 as _psycopg  # type: ignore[no-redef]
        _PSYCOPG_VERSION = 2
    except ImportError:
        _psycopg = None
        _PSYCOPG_VERSION = 0

# ---------------------------------------------------------------------------
# ENVIRONNEMENTS
# Charges depuis input_data/SQL/environments.ini. Une section par couple
# "<ENV>:<base>" (ex. [QUALIF:mdg], [PROD:dwh]) avec host/port/sid.
# Pour ajouter un environnement ou une base, editer environments.ini ;
# aucune modification de code n'est requise.
# ---------------------------------------------------------------------------
_ENV_FILE_NAME = "environments.ini"

# Fallback minimal si environments.ini est introuvable : on conserve la
# configuration historique pour ne pas casser un lancement isole du script.
_DEFAULT_ENVIRONMENTS = {
    "QUALIF": {
        "mdg": {"host": "bdd-S4PSYY00.alias", "port": "1521", "sid": "S4PSYY00"},
        "dwh": {"host": "bdd-X0PSYY02.alias", "port": "1521", "sid": "X0PSYY02"},
    },
    "PROD": {
        "mdg": {"host": "bdd-X0PSYY00.alias", "port": "1521", "sid": "X0PSYY00"},
        "dwh": {"host": "bdd-X0PSYY02.alias", "port": "1521", "sid": "X0PSYY02"},
    },
}


# Cles requises pour chaque type de base reconnu. Permet une validation au
# chargement de environments.ini, avant toute tentative de connexion.
_REQUIRED_DB_KEYS: dict[str, tuple[str, ...]] = {
    "mdg":  ("host", "port", "sid"),
    "dwh":  ("host", "port", "sid"),
    "iehe": ("host", "port", "dbname"),
}


def _load_environments(ini_path: Path) -> tuple[dict, set[str]]:
    """Lit environments.ini et renvoie (ENVIRONMENTS, VALID_DBS).

    Chaque section "<ENV>:<base>" doit fournir les cles requises selon
    le type de base (cf. _REQUIRED_DB_KEYS) :
      - Oracle (mdg, dwh) : host, port, sid
      - PostgreSQL (iehe) : host, port, dbname

    Toutes les autres cles presentes dans la section sont conservees
    telles quelles dans le dict retourne. Les sections mal formees sont
    ignorees avec un message d'alerte. Si le fichier est absent ou vide,
    on retombe sur la configuration historique en dur (compatibilite
    ascendante).
    """
    if not ini_path.is_file():
        print(f"[INFO] {ini_path.name} introuvable, configuration par defaut utilisee.")
        envs = {k: {b: dict(v) for b, v in bases.items()} for k, bases in _DEFAULT_ENVIRONMENTS.items()}
        return envs, {b for bases in envs.values() for b in bases}

    parser = configparser.ConfigParser()
    parser.read(ini_path, encoding="utf-8")

    envs: dict[str, dict[str, dict[str, str]]] = {}
    for section in parser.sections():
        if ":" not in section:
            print(f"[ATTENTION] {ini_path.name} : section '{section}' ignoree (format attendu '<ENV>:<base>').")
            continue
        env_name, db_name = section.split(":", 1)
        env_name = env_name.strip()
        db_name = db_name.strip().lower()

        # On capture toutes les cles non vides de la section (le type de
        # base determine quelles cles sont obligatoires).
        opts = {k: v.strip() for k, v in parser.items(section) if v and v.strip()}

        required = _REQUIRED_DB_KEYS.get(db_name)
        if required is None:
            print(
                f"[ATTENTION] {ini_path.name} : type de base '{db_name}' inconnu "
                f"dans la section '{section}' ; ignoree."
            )
            continue
        missing = [k for k in required if not opts.get(k)]
        if missing:
            print(
                f"[ATTENTION] {ini_path.name} : section '{section}' ignoree "
                f"(cles requises manquantes : {missing})."
            )
            continue

        envs.setdefault(env_name, {})[db_name] = opts

    if not envs:
        print(f"[INFO] {ini_path.name} vide ou invalide, configuration par defaut utilisee.")
        envs = {k: {b: dict(v) for b, v in bases.items()} for k, bases in _DEFAULT_ENVIRONMENTS.items()}

    valid_dbs = {b for bases in envs.values() for b in bases}
    return envs, valid_dbs


# ---------------------------------------------------------------------------
# DISPATCH PAR TYPE DE BASE (Oracle MDG/DWH ou PostgreSQL IEHE)
# ---------------------------------------------------------------------------
def make_conn_params(db_key: str, db_info: dict[str, str]):
    """Retourne les parametres de connexion adaptes au type de base.

    Oracle (mdg/dwh)   -> str DSN au format host:port/SID (compatible
                          `oracledb.connect(dsn=...)`).
    PostgreSQL (iehe)  -> dict kwargs (host, port, dbname) a passer en
                          `psycopg.connect(**dict)`.
    """
    if db_key in ("mdg", "dwh"):
        return oracledb.makedsn(db_info["host"], db_info["port"], sid=db_info["sid"])
    if db_key == "iehe":
        return {
            "host": db_info["host"],
            "port": int(db_info["port"]) if str(db_info["port"]).isdigit() else db_info["port"],
            "dbname": db_info["dbname"],
        }
    raise ValueError(f"Type de base non supporte : '{db_key}'")


def connect(db_key: str, conn_params, user: str, password: str):
    """Ouvre une connexion en dispatchant sur le bon driver.

    Renvoie un objet Connection (oracledb ou psycopg) avec une API DBAPI
    suffisamment proche pour le pattern `cursor.execute() / fetchall()`.
    """
    if db_key in ("mdg", "dwh"):
        return oracledb.connect(user=user, password=password, dsn=conn_params)
    if db_key == "iehe":
        if _psycopg is None:
            raise RuntimeError(
                "Connexion IEHE impossible : la librairie 'psycopg' n'est pas "
                "installee. Executer : pip install \"psycopg[binary]\" "
                "(ou en fallback : pip install psycopg2-binary)."
            )
        return _psycopg.connect(user=user, password=password, **conn_params)
    raise ValueError(f"Type de base non supporte : '{db_key}'")


# Regex de detection des placeholders {NOM} dans le corps SQL (NOM en
# majuscules ASCII, commence par une lettre).
_PLACEHOLDER_RE = re.compile(r"\{([A-Z][A-Z0-9_]*)\}")


def _resolve_default_token(raw: str | None) -> str | None:
    """Resout les tokens speciaux dans la valeur par defaut d'un parametre.
    Cote .sql, ces tokens s'ecrivent entre crochets et sont stockes sans
    crochets par `_DEFAULT_RE` (qui consomme les `[...]`).

      - `[TODAY]`     ->  TODAY     -> date du jour au format YYYYMMDD
      - `[YESTERDAY]` ->  YESTERDAY -> date d'hier au format YYYYMMDD

    Toute autre valeur est renvoyee telle quelle (litteral). La resolution
    est faite UNE FOIS au moment de la saisie (cf. ETAPE 3bis) ; la valeur
    obtenue persiste pour toute la session, evitant tout decalage si
    l'execution traverse minuit.
    """
    if raw is None:
        return None
    s = raw.strip()
    if s == "TODAY":
        return datetime.now().strftime("%Y%m%d")
    if s == "YESTERDAY":
        return (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
    return s


def _substitute_params(sql: str, values: dict[str, str]) -> str:
    """Substitue les placeholders {NOM} par les valeurs fournies.
    Les apostrophes des valeurs sont doublees pour eviter de casser le SQL
    quand la valeur est inseree dans un littéral entoure de guillemets
    simples (responsabilite du SQL d'encadrer le placeholder par des
    apostrophes si la valeur est textuelle).
    Les placeholders sans correspondance dans `values` sont laisses tels
    quels (utile pour {DATE_SUIVI} qui est gere par le flux legacy).
    """
    if not values:
        return sql

    def repl(m: re.Match) -> str:
        name = m.group(1)
        if name in values:
            return values[name].replace("'", "''")
        return m.group(0)

    return _PLACEHOLDER_RE.sub(repl, sql)


def cancel_query(conn, db_key: str) -> None:
    """Annule une requete en cours sur la connexion donnee, sans crasher
    si l'annulation echoue (logs uniquement). Utilise par le Timer de
    timeout dans execute_query.

    - oracledb : conn.cancel() (existant).
    - psycopg v3 : conn.cancel_safe() prefere (thread-safe) ; fallback
                   conn.cancel() si non disponible (anciennes versions).
    - psycopg v2 : conn.cancel().
    """
    try:
        if db_key == "iehe":
            cancel_fn = getattr(conn, "cancel_safe", None) or getattr(conn, "cancel", None)
        else:
            cancel_fn = getattr(conn, "cancel", None)
        if cancel_fn is None:
            log(f"ATTENTION : aucune methode cancel disponible pour db_key='{db_key}'.")
            return
        cancel_fn()
    except Exception as exc:
        log(f"ATTENTION : echec de l'annulation ({db_key}) : {exc}")


SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR if (SCRIPT_DIR / "input_data").exists() else SCRIPT_DIR.parent

# ---------------------------------------------------------------------------
# CHEMINS
# ---------------------------------------------------------------------------
SQL_INPUT_DIR = BASE_DIR / "input_data" / "SQL"
OUTPUT_ROOT = BASE_DIR / "output" / "SQL"

ENVIRONMENTS, VALID_DBS = _load_environments(SQL_INPUT_DIR / _ENV_FILE_NAME)

PREFIX_TO_FREQ = {"J_": "quotidien", "S_": "hebdo"}


def log(msg: str) -> None:
    """Affiche un message horodaté."""
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"  [{ts}] {msg}", flush=True)


# ---------------------------------------------------------------------------
# POST-TRAITEMENTS
# Fonctions appliquées automatiquement sur le DataFrame d'une requête
# (ajout de sous-totaux, etc.) avant l'export Excel.
# ---------------------------------------------------------------------------
def _col_to_numeric(df: pd.DataFrame, col: str) -> pd.Series:
    """Convertit une colonne (éventuellement stockée en TO_CHAR côté SQL)
    en valeurs numériques, en traitant '0'/NaN/vides comme 0."""
    return pd.to_numeric(df[col], errors="coerce").fillna(0)


def postprocess_sum_total_doublons(df: pd.DataFrame) -> pd.DataFrame:
    """Ajoute une ligne 'TOTAL_GENERAL' avec la somme de toutes les
    colonnes numériques (requête Attente / Doublons DWH)."""
    if df is None or df.empty:
        return df

    id_cols = list(df.columns[:2])
    num_cols = [c for c in df.columns if c not in id_cols]

    total_row: dict[str, object] = {id_cols[0]: "TOTAL_GENERAL"}
    if len(id_cols) > 1:
        total_row[id_cols[1]] = ""

    for c in num_cols:
        total_row[c] = int(_col_to_numeric(df, c).sum())

    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)


def postprocess_sum_boite_prefixes_stock(df: pd.DataFrame) -> pd.DataFrame:
    """Ajoute trois lignes sous-totaux (DSI*, GES*, VDC*) sommant la
    colonne 'Total' (requête Stock / Rejets par boîte DWH)."""
    if df is None or df.empty:
        return df

    boite_col = df.columns[0]
    total_col = next((c for c in df.columns if str(c).strip().lower() == "total"), None)
    if total_col is None:
        log("ATTENTION : colonne 'Total' introuvable, post-traitement ignore")
        return df

    total_num = _col_to_numeric(df, total_col)
    boites = df[boite_col].astype(str)

    rows: list[dict[str, object]] = []
    for prefix in ("DSI", "GES", "VDC"):
        mask = boites.str.startswith(prefix)
        somme = int(total_num[mask].sum())
        row = {c: "" for c in df.columns}
        row[boite_col] = f"TOTAL_{prefix}*"
        row[total_col] = somme
        rows.append(row)

    return pd.concat([df, pd.DataFrame(rows)], ignore_index=True)


def postprocess_pivot_prestations_par_offre(df: pd.DataFrame) -> pd.DataFrame:
    """Enrichit le DataFrame brut MDG_Prestations_par_offre avec les colonnes
    `_nature` et `_societe` (lookup via config/mapping_offres.yml). Le
    classeur multi-onglets aligné sur la maquette Laurence est construit
    APRÈS l'étape 4 par `build_prestations_par_offre_workbook()`.
    """
    if df is None or df.empty:
        return df
    try:
        mapping = ppo.load_mapping()
    except Exception as exc:  # noqa: BLE001
        log(f"POST-TRT pivot_prestations_par_offre : mapping KO ({exc})")
        return df
    return ppo.enrich_dataframe(df, mapping)


POST_PROCESSORS = {
    "sum_total_doublons": postprocess_sum_total_doublons,
    "sum_boite_prefixes_stock": postprocess_sum_boite_prefixes_stock,
    "pivot_prestations_par_offre": postprocess_pivot_prestations_par_offre,
}


def build_prestations_par_offre_workbook(
    results: dict[str, pd.DataFrame],
    output_dir: Path,
    heure_clean: str,
) -> Path | None:
    """Construit le classeur pivot multi-onglets pour MDG_Prestations_par_offre.

    Concatène TOUS les DataFrames issus de la catégorie (quotidien + hebdo si
    présents), puis appelle `prestations_par_offre_lib.build_pivot_workbook`.
    Retourne le chemin du fichier produit, ou None si aucun DataFrame de la
    catégorie n'est présent.
    """
    relevant = [
        df for name, df in results.items()
        if df is not None and not df.empty
        and "_nature" in df.columns and "_societe" in df.columns
    ]
    if not relevant:
        return None
    try:
        mapping = ppo.load_mapping()
    except Exception as exc:  # noqa: BLE001
        log(f"BUILDER Prestations_Par_Offre : mapping KO ({exc})")
        return None
    df_concat = pd.concat(relevant, ignore_index=True)
    out_path = output_dir / (
        f"MDG_Prestations_par_offre_{heure_clean}_Pivot.xlsx")
    try:
        ppo.build_pivot_workbook(df_concat, out_path, mapping)
    except Exception as exc:  # noqa: BLE001
        log(f"BUILDER Prestations_Par_Offre : ECHEC construction ({exc})")
        return None
    log(f"BUILDER Prestations_Par_Offre : {out_path.name} ecrit "
        f"({len(df_concat)} lignes source)")
    return out_path


# ---------------------------------------------------------------------------
# MISE A JOUR DU CLASSEUR 00-PREST_<ANNEE>.xlsx
# Apres execution, si les deux requetes MDG_Prestations + DWH ont produit
# leurs resultats pour la date saisie, on remplit le bloc journalier
# correspondant dans 00-PREST_<ANNEE>.xlsx (idempotent : reecriture du bloc
# si une entree existe deja pour cette date).
# ---------------------------------------------------------------------------
PREST_MDG_SHEET = "00_MDG_IT_Prest_V3"
PREST_DWH_SHEET = "00_IT_DWH_Prestations"

# Ordre des libelles dans le bloc (10 lignes apres l'en-tete).
PREST_ROW_LABELS = [
    "MDG_Intégrés", "MDG_Payés", "MDG_Relevés", "MDG_Attente", "MDG_Rejets",
    "DWH-Intégré", "DWH_Payé", "DWH_Relevé", "DWH-Attente", "DWH_Rejets",
]

# Schema des en-tetes de colonne (17 colonnes : A=libelle + 16 valeurs).
# Doit correspondre au schema en vigueur dans les feuilles 042026+.
PREST_HEADERS = [
    "DRASS", "NMFOU", "NMASS", "LUMAS", "NMCMU",
    "TPVIA", "TPHOS", "TPCVM", "OXA", "Total",
    "ASSUR", "SELFC", "TIERS", "REGUL", "Total", "Total Général",
]

# Mapping nom de colonne attendu cote SQL -> indice colonne Excel (B=2 ...).
# Les noms cote DataFrame proviennent de la requete (alias SQL) ; on tolere
# la casse via une normalisation upper().
PREST_SQL_TO_COL = {
    "DRASS": 2, "NMFOU": 3, "NMASS": 4, "LUMAS": 5, "NMCMU": 6,
    "TPVIA": 7, "TPHOS": 8, "TPCVM": 9, "OXA": 10,
    "TOTAL_AUTO": 11,
    "ASSUR": 12, "SELFC": 13, "TIERS": 14, "REGUL": 15,
    "TOTAL_MANUEL": 16,
}

# Colonnes recevant une formule de ratio en ligne n+11 (rejets/(integres-attente)).
# (col_letter, row_offset_rejets, row_offset_integres, row_offset_attente)
# Offsets relatifs au numero de ligne de l'en-tete du bloc (start_row).
# Convention : start_row = en-tete, start_row+1..+5 = MDG, +6..+10 = DWH, +11 = ratios.
PREST_RATIO_FORMULAS = [
    # Cols D..J + Q : ratio "MDG" (rejets MDG / (integres MDG - attente MDG)).
    ("D", 5, 1, 4), ("E", 5, 1, 4), ("F", 5, 1, 4),
    ("G", 5, 1, 4), ("H", 5, 1, 4), ("I", 5, 1, 4), ("J", 5, 1, 4),
    ("Q", 5, 1, 4),
    # Col M (SELFC) : ratio "DWH" (rejets DWH / (integres DWH - attente DWH)).
    ("M", 10, 6, 9),
]


def _normalize_prest_df(df: pd.DataFrame) -> pd.DataFrame:
    """Normalise les noms de colonnes de la requete (UPPER) pour faciliter
    le mapping vers les colonnes Excel."""
    df = df.copy()
    df.columns = [str(c).upper() for c in df.columns]
    return df


def _prest_value(df: pd.DataFrame, requete_label: str, col_name: str):
    """Recupere la valeur de la colonne `col_name` pour la ligne dont la
    colonne `REQUETE` vaut `requete_label`. Retourne 0 si introuvable."""
    if "REQUETE" not in df.columns or col_name not in df.columns:
        return 0
    rows = df[df["REQUETE"] == requete_label]
    if rows.empty:
        return 0
    val = rows.iloc[0][col_name]
    try:
        return int(val) if pd.notna(val) else 0
    except (TypeError, ValueError):
        return val if pd.notna(val) else 0


def _find_block_for_date(ws, date_obj: datetime) -> int | None:
    """Cherche dans la feuille un bloc dont la cellule A (ligne d'en-tete)
    contient la date donnee. Renvoie le numero de ligne de l'en-tete, ou
    None si aucun bloc trouve.

    Parcours en remontant depuis ws.max_row (les ajouts recents sont en
    bas) et gere les deux types possibles renvoyes par openpyxl :
    `datetime.datetime` (cas usuel) et `datetime.date` pour lequel
    `hasattr(v, "date")` est faux.
    """
    target = date_obj.date()
    for r in range(ws.max_row, 0, -1):
        v = ws.cell(row=r, column=1).value
        v_date = v.date() if hasattr(v, "date") else v
        if v_date == target:
            return r
    return None


def _next_block_row(ws) -> int:
    """Determine la ligne d'en-tete pour un nouveau bloc, juste apres le
    dernier contenu non vide de la feuille. Saute une ligne pour aerer.

    Parcours en remontant depuis ws.max_row pour s'arreter au premier
    contenu rencontre (plus rapide sur feuille volumineuse).
    """
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, 18)):
            return r + 2
    return 1


def _write_prest_headers(ws, start_row: int, date_obj: datetime) -> None:
    """Ecrit la ligne d'en-tete (date en A + 16 libelles en B..Q)."""
    ws.cell(row=start_row, column=1, value=date_obj)
    for i, h in enumerate(PREST_HEADERS, start=2):
        ws.cell(row=start_row, column=i, value=h)


def _write_prest_block(
    ws, start_row: int, date_obj: datetime,
    df_mdg: pd.DataFrame, df_dwh: pd.DataFrame,
) -> None:
    """Ecrit (ou ecrase) un bloc journalier complet a partir de start_row.

    Layout (17 colonnes, A..Q) :
      start_row     : date + en-tetes
      start_row+1.5 : 5 lignes MDG_*
      start_row+6.10: 5 lignes DWH-*
      start_row+11  : ligne de ratios (formules)
      start_row+12  : ligne vide (aeration)
    """
    _write_prest_headers(ws, start_row, date_obj)

    df_mdg_n = _normalize_prest_df(df_mdg)
    df_dwh_n = _normalize_prest_df(df_dwh)

    for offset, label in enumerate(PREST_ROW_LABELS, start=1):
        row = start_row + offset
        df = df_mdg_n if offset <= 5 else df_dwh_n
        ws.cell(row=row, column=1, value=label)
        for sql_col, xls_col in PREST_SQL_TO_COL.items():
            ws.cell(row=row, column=xls_col, value=_prest_value(df, label, sql_col))
        # Q (Total Général) = K (Total_Auto) + P (Total_Manuel) en formule.
        ws.cell(row=row, column=17, value=f"=K{row}+P{row}")

    # Ligne de ratios (start_row + 11).
    ratio_row = start_row + 11
    for col_letter, off_rej, off_int, off_att in PREST_RATIO_FORMULAS:
        r_rej = start_row + off_rej
        r_int = start_row + off_int
        r_att = start_row + off_att
        ws.cell(
            row=ratio_row, column=ord(col_letter) - 64,
            value=f"=IFERROR({col_letter}{r_rej}/({col_letter}{r_int}-{col_letter}{r_att}), 0)",
        )

    # Ligne start_row+12 laissee vide volontairement (aeration).


def update_prest_workbook(
    results: dict[str, pd.DataFrame],
    quotidien_date: str | None,
    workbook_dir: Path,
) -> Path | None:
    """Met a jour 00-PREST_<ANNEE>.xlsx avec les resultats des requetes
    MDG_Prestations + DWH du jour.

    Conditions :
      - Les deux DataFrames `PREST_MDG_SHEET` et `PREST_DWH_SHEET` doivent
        etre presents dans `results` (sinon retourne None).
      - Une date quotidienne doit avoir ete saisie (sinon retourne None).
      - Le classeur 00-PREST_<ANNEE>.xlsx doit exister dans `workbook_dir`
        (typiquement output/SQL/).

    Comportement :
      - Si un bloc pour la date existe deja dans la feuille MMYYYY,
        il est ecrase (idempotent).
      - Sinon un nouveau bloc 13 lignes est ajoute apres le dernier bloc.
      - La feuille MMYYYY est creee si absente.

    Retourne le chemin du fichier mis a jour, ou None si rien n'a ete fait.
    """
    if PREST_MDG_SHEET not in results or PREST_DWH_SHEET not in results:
        return None
    if not quotidien_date:
        return None

    df_mdg = results[PREST_MDG_SHEET]
    df_dwh = results[PREST_DWH_SHEET]
    if df_mdg is None or df_dwh is None or df_mdg.empty or df_dwh.empty:
        log("ATTENTION : DataFrames MDG/DWH vides, mise a jour 00-PREST ignoree.")
        return None

    try:
        date_obj = datetime.strptime(quotidien_date, "%d/%m/%Y")
    except ValueError:
        log(f"ATTENTION : date '{quotidien_date}' invalide, mise a jour 00-PREST ignoree.")
        return None

    workbook_path = workbook_dir / f"00-PREST_{date_obj.year}.xlsx"
    if not workbook_path.exists():
        log(
            f"ATTENTION : {workbook_path} introuvable, "
            "mise a jour 00-PREST ignoree."
        )
        return None

    # Import local pour ne pas dependre d'openpyxl si la feature n'est pas utilisee.
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path)
    sheet_name = date_obj.strftime("%m%Y")
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        log(f"00-PREST : feuille '{sheet_name}' creee.")
    else:
        ws = wb[sheet_name]

    existing = _find_block_for_date(ws, date_obj)
    if existing is not None:
        start_row = existing
        log(f"00-PREST : bloc existant pour {quotidien_date} ligne {start_row} -> reecriture.")
    else:
        start_row = _next_block_row(ws)
        log(f"00-PREST : nouveau bloc ajoute ligne {start_row} pour {quotidien_date}.")

    _write_prest_block(ws, start_row, date_obj, df_mdg, df_dwh)
    try:
        wb.save(workbook_path)
    except PermissionError:
        log(
            f"ERREUR : impossible d'enregistrer {workbook_path.name}. "
            "Verifiez qu'il n'est pas ouvert dans Excel."
        )
        return None
    log(f"00-PREST : fichier sauvegarde -> {workbook_path}")
    return workbook_path


# ---------------------------------------------------------------------------
# CHARGEMENT DES REQUÊTES DEPUIS LE DOSSIER input_data/SQL
# ---------------------------------------------------------------------------
_META_RE = re.compile(r"^\s*--\s*([A-Za-z_][A-Za-z0-9_]*)\s*:\s*(.+?)\s*$")
# Detecte les declarations de parametres : `-- param.NOM: prompt [defaut]`.
# NOM = majuscules ASCII + chiffres + `_`, commence par une lettre.
_PARAM_RE = re.compile(
    r"^\s*--\s*param\.([A-Z][A-Z0-9_]*)\s*:\s*(.+?)\s*$"
)
# Capture le defaut entre crochets en fin de ligne du prompt.
_DEFAULT_RE = re.compile(r"^(?P<prompt>.+?)\s*\[(?P<default>[^\]]+)\]\s*$")


def _parse_sql_file(path: Path) -> dict:
    """Lit un fichier .sql et en extrait :
      - meta            : metadonnees globales `-- key: value` (db, post_process, ...)
      - params          : dict {NOM: {"prompt": str, "default": Optional[str]}}
      - params_order    : ordre de declaration des parametres
      - sql             : corps SQL

    La phase de lecture des en-tetes s'arrete a la premiere ligne non-vide
    qui n'est ni un commentaire `-- key: value`, ni un `-- param.X:`.
    """
    text = path.read_text(encoding="utf-8-sig")
    lines = text.splitlines()

    meta: dict[str, str] = {}
    params: dict[str, dict[str, str | None]] = {}
    params_order: list[str] = []
    i = 0
    while i < len(lines):
        raw = lines[i]
        stripped = raw.strip()
        if stripped == "":
            i += 1
            continue
        m_param = _PARAM_RE.match(raw)
        if m_param is not None:
            name = m_param.group(1)
            rest = m_param.group(2).strip()
            md = _DEFAULT_RE.match(rest)
            if md:
                prompt = md.group("prompt").strip()
                default = md.group("default").strip()
            else:
                prompt = rest
                default = None
            if name not in params:
                params[name] = {"prompt": prompt, "default": default}
                params_order.append(name)
            i += 1
            continue
        m = _META_RE.match(raw)
        if m is None:
            break
        meta[m.group(1).lower()] = m.group(2).strip()
        i += 1

    sql_body = "\n".join(lines[i:]).strip()
    # Retire un eventuel ';' final : oracledb (et psycopg) refusent un
    # delimiteur de fin de requete (`;` ou `/`) qui n'est valide qu'en
    # SQL*Plus. On laisse le .sql lisible cote utilisateur tout en
    # acceptant les conventions habituelles.
    while sql_body.endswith((";", "/")):
        sql_body = sql_body[:-1].rstrip()
    return {
        "meta": meta,
        "params": params,
        "params_order": params_order,
        "sql": sql_body,
    }


def _detect_freq_from_name(filename: str) -> str | None:
    """Retourne la frequence ('quotidien' ou 'hebdo') deduite du prefixe
    du nom de fichier (`J_` ou `S_`), ou None si le fichier ne suit pas
    la convention.
    """
    for prefix, freq in PREFIX_TO_FREQ.items():
        if filename.startswith(prefix):
            return freq
    return None


def load_queries(sql_dir: Path) -> dict[str, dict]:
    """Scanne sql_dir (recursivement) et retourne un dict
    { nom_feuille: {freq, db, sql, post_process?, category} } dans l'ordre
    alphabetique (ordre reproductible des feuilles Excel).

    Convention de nommage : seuls les fichiers prefixes par `J_` (quotidien)
    ou `S_` (hebdo) sont auto-executes. Tout autre `.sql` est ignore.
    Le nom de feuille Excel correspond au nom de fichier SANS le prefixe.

    Categorisation : chaque .sql est rattache a une categorie deduite de
    son sous-dossier sous input_data/SQL/. Un fichier place directement
    a la racine de input_data/SQL/ recoit la categorie "Autre". Les
    sous-dossiers techniques (DDL) sont ignores.
    """
    if not sql_dir.is_dir():
        print(f"[ERREUR] Dossier SQL introuvable : {sql_dir}")
        sys.exit(1)

    excluded_subdirs = {"DDL"}
    all_files: list[Path] = []
    for p in sorted(sql_dir.rglob("*.sql")):
        rel_parts = p.relative_to(sql_dir).parts
        if rel_parts[0] in excluded_subdirs and len(rel_parts) > 1:
            continue
        all_files.append(p)

    if not all_files:
        print(f"[ERREUR] Aucun fichier .sql trouve sous : {sql_dir}")
        sys.exit(1)

    files_with_freq: list[tuple[Path, str]] = []
    skipped: list[Path] = []
    for p in all_files:
        freq = _detect_freq_from_name(p.name)
        if freq is None:
            skipped.append(p)
        else:
            files_with_freq.append((p, freq))

    if skipped:
        print(f"  -> {len(skipped)} requete(s) ignoree(s) (sans prefixe 'J_' ou 'S_') :")
        for p in skipped:
            print(f"     - {p.relative_to(sql_dir)}")

    if not files_with_freq:
        print(
            f"[ERREUR] Aucun fichier .sql avec prefixe 'J_' ou 'S_' sous : {sql_dir}"
        )
        sys.exit(1)

    queries: dict[str, dict] = {}
    erreurs: list[str] = []

    for path, freq in files_with_freq:
        name = path.stem[2:]
        parsed = _parse_sql_file(path)
        meta = parsed["meta"]
        sql_body = parsed["sql"]

        if not sql_body:
            erreurs.append(f"{path.relative_to(sql_dir)} : corps SQL vide")
            continue

        db = meta.get("db", "").lower()
        if db not in VALID_DBS:
            erreurs.append(
                f"{path.relative_to(sql_dir)} : metadonnee 'db' manquante ou invalide "
                f"(attendu {sorted(VALID_DBS)}, recu '{db}')"
            )
            continue

        rel_parts = path.relative_to(sql_dir).parts
        category = rel_parts[0] if len(rel_parts) > 1 else "Autre"

        # --- Validation des placeholders {NOM} vs declarations -- param.NOM: ---
        # `DATE_SUIVI` est tolere sans declaration (backward compat : il est
        # automatiquement substitue par la date de frequence saisie au
        # demarrage).
        placeholders = set(_PLACEHOLDER_RE.findall(sql_body))
        declared = set(parsed["params"].keys())
        legacy = {"DATE_SUIVI"}
        missing = sorted(placeholders - declared - legacy)
        if missing:
            erreurs.append(
                f"{path.relative_to(sql_dir)} : placeholders {missing} utilises "
                f"dans le SQL mais non declares en en-tete (-- param.<NOM>: ...)."
            )
            continue
        unused = sorted(declared - placeholders)
        for u in unused:
            log(f"ATTENTION {path.name} : param.{u} declare mais inutilise dans le SQL.")

        entry: dict = {
            "freq": freq, "db": db, "sql": sql_body, "category": category,
            "params": parsed["params"], "params_order": parsed["params_order"],
        }

        pp = meta.get("post_process")
        if pp:
            if pp not in POST_PROCESSORS:
                log(f"ATTENTION {path.name} : post_process '{pp}' inconnu - ignore")
            else:
                entry["post_process"] = pp

        queries[name] = entry

    if erreurs:
        print("[ERREUR] Probleme(s) de metadonnees dans input_data/SQL :")
        for e in erreurs:
            print(f"  - {e}")
        sys.exit(1)

    return queries


# ---------------------------------------------------------------------------
# EXECUTION (Oracle MDG/DWH ou PostgreSQL IEHE — dispatch sur db_key)
# ---------------------------------------------------------------------------
def _is_timeout_error(exc: Exception, db_key: str) -> bool:
    """Detecte si une exception remontee par le driver correspond a une
    annulation de requete (cas timeout). Specifique a chaque driver :
      - Oracle : code 'ORA-01013' dans le message.
      - psycopg v3 / v2 : exception `QueryCanceled` (sous-classe d'`OperationalError`).
    """
    if db_key in ("mdg", "dwh"):
        return "ORA-01013" in str(exc)
    if db_key == "iehe":
        return exc.__class__.__name__ in ("QueryCanceled", "QueryCanceledError")
    return False


def execute_query(
    db_key: str, conn_params, user: str, password: str,
    sheet_name: str, query: str, timeout: int,
) -> tuple[str, pd.DataFrame | None, str]:
    """Execute une requete dans sa propre connexion (pour le parallelisme).
    Retourne (sheet_name, DataFrame ou None, message de log).

    `conn_params` est soit un DSN string (Oracle) soit un dict kwargs
    (PostgreSQL). Le dispatch est gere par `connect()` et `cancel_query()`.
    """
    t0 = time.time()
    log(f"DEBUT   {sheet_name}")
    try:
        conn = connect(db_key, conn_params, user, password)
        cursor = conn.cursor()
        timer = threading.Timer(timeout, lambda: cancel_query(conn, db_key))
        timer.start()
        try:
            cursor.execute(query)
            cols = [c[0] for c in cursor.description] if cursor.description else []
            rows = cursor.fetchall()
        finally:
            timer.cancel()
            cursor.close()
        conn.close()
        df = pd.DataFrame(rows, columns=cols)
        duree = time.time() - t0
        msg = f"OK      {sheet_name} -> {len(df)} lignes en {duree:.1f}s"
        log(msg)
        return sheet_name, df, msg
    except Exception as exc:
        duree = time.time() - t0
        if _is_timeout_error(exc, db_key):
            msg = f"TIMEOUT {sheet_name} ({timeout}s) - requete annulee [{duree:.1f}s]"
        else:
            msg = f"ERREUR  {sheet_name} : {exc} [{duree:.1f}s]"
        log(msg)
        return sheet_name, None, msg


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main() -> None:
    print("=" * 60)
    print("       EXECUTION REQUETES SQL ORACLE  (V2)")
    print("=" * 60)
    print(f"  Source  SQL : {SQL_INPUT_DIR}")
    print(f"  Dossier out : {OUTPUT_ROOT}")

    # =================================================================
    # ETAPE 0 : Chargement des requetes depuis input_data/SQL/
    # Convention : 'J_' = quotidien, 'S_' = hebdo, autres = ignores
    # =================================================================
    QUERIES = load_queries(SQL_INPUT_DIR)
    nb_j = sum(1 for v in QUERIES.values() if v["freq"] == "quotidien")
    nb_s = sum(1 for v in QUERIES.values() if v["freq"] == "hebdo")
    print(
        f"\n  -> {len(QUERIES)} requete(s) detectee(s) "
        f"({nb_j} quotidienne(s) 'J_', {nb_s} hebdo 'S_')"
    )

    # =================================================================
    # ETAPE 1 : Selection des categories a executer (multi-selection)
    # Une categorie = un sous-dossier de input_data/SQL/ (ex. KPI5_Rejets_Hebdo,
    # MDG_Prestations, DWH...). Multi-selection par numeros separes par
    # des virgules, ou 'A' pour tout. La saisie est insensible aux espaces.
    # =================================================================
    print("\n--- ETAPE 1 : Quelles categories executer ? ---\n")
    categories = sorted({v["category"] for v in QUERIES.values()})
    cat_counts = {c: sum(1 for v in QUERIES.values() if v["category"] == c) for c in categories}
    cat_freqs = {
        c: sorted({v["freq"] for v in QUERIES.values() if v["category"] == c})
        for c in categories
    }
    for i, cat in enumerate(categories, 1):
        freqs = "+".join("J" if f == "quotidien" else "S" for f in cat_freqs[cat])
        print(f"  {i} - {cat:30s} [{cat_counts[cat]:>2} requete(s) | {freqs}]")
    print(f"  A - Toutes les categories         [{len(QUERIES):>2} requete(s)]")
    choix_cat = input("\n  Votre choix (ex: 1,3 ou A) : ").strip().upper()

    if choix_cat == "A" or choix_cat == "":
        cats_selected = set(categories)
    else:
        cats_selected = set()
        for token in choix_cat.split(","):
            token = token.strip()
            if not token.isdigit():
                print(f"[ERREUR] Choix invalide : '{token}'.")
                sys.exit(1)
            idx = int(token) - 1
            if not (0 <= idx < len(categories)):
                print(f"[ERREUR] Numero hors plage : '{token}'.")
                sys.exit(1)
            cats_selected.add(categories[idx])

    selected = {k: v for k, v in QUERIES.items() if v["category"] in cats_selected}
    if not selected:
        print("[ERREUR] Aucune requete correspondante aux categories choisies.")
        sys.exit(1)

    print(f"\n  -> {len(selected)} requete(s) selectionnee(s) "
          f"dans {len(cats_selected)} categorie(s) : {', '.join(sorted(cats_selected))}")

    # Valeurs des parametres utilisateur (remplies par l'etape 3bis dediee
    # si au moins une requete declare des `-- param.X:` au-dela du legacy
    # {DATE_SUIVI}). Reste vide sinon : la substitution est un no-op.
    param_values: dict[str, str] = {}

    # =================================================================
    # ETAPE 2 : Selection de l'environnement (PROD ou QUALIF)
    # =================================================================
    print("\n--- ETAPE 2 : Environnement ---\n")
    env_keys = list(ENVIRONMENTS.keys())
    for i, env_name in enumerate(env_keys, 1):
        print(f"  {i} - {env_name}")
    choix_env = input(f"\n  Votre choix (1-{len(env_keys)}) : ").strip()
    try:
        env = ENVIRONMENTS[env_keys[int(choix_env) - 1]]
        env_nom = env_keys[int(choix_env) - 1]
    except (ValueError, IndexError):
        print("[ERREUR] Choix invalide.")
        sys.exit(1)

    print(f"\n  -> Environnement : {env_nom}")
    for db_key, info in sorted(env.items()):
        label = db_key.upper()
        # Affichage adapte selon le type de base (sid pour Oracle, dbname pour Postgres).
        ident = info.get("sid") or info.get("dbname") or "?"
        print(f"     {label:5s}: {info['host']} / {ident}")

    # =================================================================
    # ETAPE 3 : Informations de connexion
    # Le .bat parent peut transmettre les identifiants Oracle via les
    # variables d'environnement ORA_USER / ORA_PASSWORD et, optionnellement,
    # les identifiants IEHE (PostgreSQL) via IEHE_USER / IEHE_PASSWORD.
    # Si IEHE_* est absent, on reutilise les creds Oracle.
    # Fallback interactif (mode standalone sans .bat) en l'absence d'env vars.
    # =================================================================
    print("\n--- ETAPE 3 : Connexion ---\n")
    ora_user = (os.environ.get("ORA_USER") or "").strip()
    ora_password = os.environ.get("ORA_PASSWORD") or ""
    if ora_user:
        print(f"  Nom utilisateur Oracle : {ora_user}  (depuis ORA_USER)")
    else:
        ora_user = input("  Nom utilisateur Oracle : ").strip()
    if ora_password:
        print("  Mot de passe Oracle    : ********  (depuis ORA_PASSWORD)")
    else:
        ora_password = getpass.getpass("  Mot de passe Oracle    : ")
    if not ora_user or not ora_password:
        print("[ERREUR] Identifiants Oracle vides.")
        sys.exit(1)

    # Credentials IEHE : par defaut on reutilise les creds Oracle.
    # Override via IEHE_USER / IEHE_PASSWORD si les comptes different.
    iehe_user = (os.environ.get("IEHE_USER") or "").strip() or ora_user
    iehe_password = os.environ.get("IEHE_PASSWORD") or ora_password
    if (os.environ.get("IEHE_USER") or "").strip():
        print(f"  Nom utilisateur IEHE   : {iehe_user}  (depuis IEHE_USER)")
    if os.environ.get("IEHE_PASSWORD"):
        print("  Mot de passe IEHE      : ********  (depuis IEHE_PASSWORD)")

    # Aliases historiques (utilises par le reste du flux pour les requetes Oracle).
    user = ora_user
    password = ora_password

    # --- Dates de suivi (une par frequence presente) ---
    freqs_presentes = {v["freq"] for v in selected.values()}
    dates: dict[str, str] = {}

    if "quotidien" in freqs_presentes:
        d = input("\n  Date QUOTIDIENNE (DD/MM/YYYY) : ").strip()
        if not re.match(r"^\d{2}/\d{2}/\d{4}$", d):
            print(f"[ERREUR] Format invalide : '{d}'")
            sys.exit(1)
        dates["quotidien"] = d

    if "hebdo" in freqs_presentes:
        d = input("  Date HEBDOMADAIRE (DD/MM/YYYY) : ").strip()
        if not re.match(r"^\d{2}/\d{2}/\d{4}$", d):
            print(f"[ERREUR] Format invalide : '{d}'")
            sys.exit(1)
        dates["hebdo"] = d

    timeout_str = input("  Timeout par requete en secondes (defaut 300) : ").strip()
    timeout = int(timeout_str) if timeout_str else 300

    # =================================================================
    # ETAPE 3bis : Saisie des parametres utilisateur (-- param.X:)
    # On agrege l'union des parametres declares par les requetes
    # selectionnees (hors DATE_SUIVI, gere par le flux par frequence).
    # Chaque parametre est demande UNE FOIS au demarrage ; la valeur
    # saisie est appliquee a toutes les requetes qui le referencent.
    # =================================================================
    # Construction ordonnee de l'union des parametres.
    all_params: dict[str, dict[str, str | None]] = {}
    for q in selected.values():
        for p_name in q.get("params_order", []):
            if p_name == "DATE_SUIVI":
                # DATE_SUIVI est gere par les dates de frequence ; on ignore
                # toute declaration explicite (le legacy l'emporte).
                continue
            if p_name not in all_params:
                all_params[p_name] = q["params"][p_name]

    if all_params:
        print("\n--- ETAPE 3bis : Parametres ---\n")
        for p_name, p_def in all_params.items():
            default_raw = p_def.get("default")
            default_resolved = _resolve_default_token(default_raw) if default_raw else None
            prompt_text = p_def.get("prompt") or p_name
            if default_resolved is not None:
                val = input(f"  {prompt_text} [{default_resolved}] : ").strip()
                if not val:
                    val = default_resolved
            else:
                val = input(f"  {prompt_text} : ").strip()
                if not val:
                    print(f"[ERREUR] Parametre obligatoire '{p_name}' non renseigne.")
                    sys.exit(1)
            param_values[p_name] = val

    # --- Construction des parametres de connexion par db_key ---
    # Pour chaque base presente dans l'env, on prepare son objet de
    # connexion (str DSN pour Oracle, dict kwargs pour PostgreSQL).
    conn_params_map: dict[str, object] = {
        db_key: make_conn_params(db_key, info) for db_key, info in env.items()
    }

    # Selectionne les creds adaptes au type de base.
    def creds_for(db_key: str) -> tuple[str, str]:
        if db_key == "iehe":
            return iehe_user, iehe_password
        return ora_user, ora_password

    # --- Test de connexion sur chaque base utilisee ---
    dbs_utilisees = {v["db"] for v in selected.values()}
    if "iehe" in dbs_utilisees and _psycopg is None:
        print(
            "[ERREUR] Une requete IEHE est selectionnee mais la librairie "
            "'psycopg' n'est pas installee. Installer avec :\n"
            "  pip install \"psycopg[binary]\"\n"
            "ou en fallback :\n"
            "  pip install psycopg2-binary"
        )
        sys.exit(1)

    for db_key in sorted(dbs_utilisees):
        if db_key not in conn_params_map:
            print(
                f"[ERREUR] Base '{db_key}' utilisee par une requete mais non "
                f"declaree dans environments.ini pour l'environnement {env_nom}."
            )
            sys.exit(1)
        label = db_key.upper()
        print(f"\n  Test de connexion {label}... ", end="", flush=True)
        try:
            test_user, test_password = creds_for(db_key)
            test_conn = connect(db_key, conn_params_map[db_key], test_user, test_password)
            test_conn.close()
            print("OK")
        except Exception as exc:
            print(f"ECHEC\n  Base {label} : {exc}")
            sys.exit(1)

    # --- Dossier de sortie : output/SQL/<date_saisie>/ ---
    # Un xlsx par categorie : <CATEGORIE>_<HHMMSS>_Resultats_SQL.xlsx
    date_prefix = dates.get("quotidien", dates.get("hebdo", ""))
    date_prefix_clean = date_prefix.replace("/", "")
    heure_clean = datetime.now().strftime("%H%M%S")
    output_dir = OUTPUT_ROOT / date_prefix_clean
    output_dir.mkdir(parents=True, exist_ok=True)

    output_files = {
        cat: output_dir / f"{cat}_{heure_clean}_Resultats_SQL.xlsx"
        for cat in cats_selected
    }

    # =================================================================
    # ETAPE 4 : Execution parallele
    # =================================================================
    print("\n--- ETAPE 4 : Execution ---\n")
    print(f"  Environnement : {env_nom}")
    for freq, d in dates.items():
        print(f"  Date {freq:12s}: {d}")
    print(f"  Timeout       : {timeout}s")
    print(f"  Requetes      : {len(selected)}")
    print(f"  Parallelisme  : {min(len(selected), 4)} threads")
    print(f"  Dossier sortie: {output_dir}")
    for cat, path in sorted(output_files.items()):
        print(f"    - {cat:30s} -> {path.name}")
    print()

    t_global = time.time()
    results: dict[str, pd.DataFrame] = {}

    with ThreadPoolExecutor(max_workers=min(len(selected), 4)) as pool:
        futures = {}
        for sheet_name, entry in selected.items():
            date_pour_requete = dates[entry["freq"]]
            # Substitution legacy {DATE_SUIVI} + parametres user (cf. ETAPE 3bis).
            query = entry["sql"].replace("{DATE_SUIVI}", date_pour_requete)
            query = _substitute_params(query, param_values)
            db_key = entry["db"]
            q_user, q_password = creds_for(db_key)
            future = pool.submit(
                execute_query,
                db_key, conn_params_map[db_key], q_user, q_password,
                sheet_name, query, timeout,
            )
            futures[future] = sheet_name

        for future in as_completed(futures):
            sheet_name, df, msg = future.result()
            if df is not None:
                pp_key = selected[sheet_name].get("post_process")
                pp_fn = POST_PROCESSORS.get(pp_key) if pp_key else None
                if pp_fn is not None:
                    try:
                        df = pp_fn(df)
                        log(f"POST-TRT {sheet_name} : {pp_key} applique")
                    except Exception as exc:
                        log(f"POST-TRT {sheet_name} : ECHEC ({exc})")
                results[sheet_name] = df

    # --- Export Excel : un fichier par categorie, dans l'ordre de selected ---
    written_files: list[Path] = []
    if results:
        for cat in sorted(cats_selected):
            sheets_for_cat = [
                s for s in selected
                if selected[s]["category"] == cat and s in results
            ]
            if not sheets_for_cat:
                continue
            path = output_files[cat]
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                for sheet_name in sheets_for_cat:
                    safe_name = sheet_name[:31]
                    results[sheet_name].to_excel(writer, sheet_name=safe_name, index=False)
            written_files.append(path)

    # =================================================================
    # ETAPE 4bis : Builder Excel pivot pour MDG_Prestations_par_offre
    # Si la categorie a produit des DataFrames enrichis (_nature + _societe
    # ajoutees par le post-processor), on construit le classeur multi-
    # onglets aligne sur la maquette Laurence (TCD1, Prestations par offre,
    # Prevoyance, Ste 010/044/052, TCD2, MEN_ORIGINE).
    # =================================================================
    ppo_workbook = build_prestations_par_offre_workbook(
        results=results, output_dir=output_dir, heure_clean=heure_clean,
    )
    if ppo_workbook is not None:
        written_files.append(ppo_workbook)

    # =================================================================
    # ETAPE 5 : Mise a jour du classeur 00-PREST_<ANNEE>.xlsx
    # Conditions : les requetes 00_MDG_IT_Prest_V3 ET 00_IT_DWH_Prestations
    # ont produit un DataFrame, et une date quotidienne a ete saisie.
    # Sinon l'etape est silencieusement sautee.
    # =================================================================
    prest_updated = update_prest_workbook(
        results=results,
        quotidien_date=dates.get("quotidien"),
        workbook_dir=OUTPUT_ROOT,
    )

    duree_totale = time.time() - t_global
    print(f"\n{'=' * 60}")
    print(f"  TERMINE en {duree_totale:.1f}s")
    print(f"  {len(results)}/{len(selected)} requete(s) reussie(s)")
    for path in written_files:
        print(f"  Fichier : {path}")
    if prest_updated is not None:
        print(f"  00-PREST : {prest_updated}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()