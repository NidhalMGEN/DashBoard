# -*- coding: utf-8 -*-
"""
09_taux_noemisation.py — Taux de noémisation par client/offre
==============================================================

Entrée  : Input_Data/{PREFIX}_Taux_Noemie*.xlsx   (PREFIX = DDMMYYYY)
          extraction Power BI brute — rapport "Taux de Noémisation".
          SEULES les colonnes libcrt / Actifs / non Actifs / Non_noémisable
          sont attendues : le regroupement client/offre, le rattachement
          Individuel/Collectif et l'ordre d'affichage ne figurent PAS dans
          l'extraction — ils sont figés dans NOEMIE_GROUPES ci-dessous.
Sortie  : Output/{PREFIX}_Taux_Noemie*.xlsx
          copie du fichier source + feuille 'Taux_Noémisation'.
          La feuille source n'est jamais modifiée.

Feuille générée (principe Feuil7 du TCD) :
  une ligne par client/offre consolidé, triée par ordre d'affichage,
  puis Total / Total Collectif / Total Individuel.

Taux = Actifs / (Actifs + non Actifs + Non_noémisable)

Les 3 indicateurs remontés au dashboard sont les totaux :
  taux total · taux Collectif · taux Individuel

Usage :
  python 09_taux_noemisation.py
"""

from __future__ import annotations

import re
import shutil
import sys
import unicodedata
import warnings
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Lancé par le pipeline, ce script hérite de PYTHONIOENCODING=utf-8
# (cf. pipeline_runner.py). Lancé directement, il hérite de la console Windows
# en cp1252, incapable d'encoder les flèches '→' des messages d'alerte : la
# trace planterait sur un UnicodeEncodeError au lieu d'afficher le diagnostic.
for _flux in (sys.stdout, sys.stderr):
    if hasattr(_flux, "reconfigure"):
        _flux.reconfigure(encoding="utf-8", errors="replace")

# ─── CONFIG FICHIER ───────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR   = SCRIPT_DIR if (SCRIPT_DIR / "Input_Data").exists() else SCRIPT_DIR.parent
INPUT_DIR  = BASE_DIR / "Input_Data"
OUTPUT_DIR = BASE_DIR / "Output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
INPUT_DIR.mkdir(parents=True, exist_ok=True)

# Nommage convenu avec PILLON Laurence : jjmmaaaa_Taux_Noémie.xlsx
# (accent et séparateur libres — "Noemie", "Noémie", "Taux Noemie"…)
NOEMIE_FILENAME_RE = re.compile(r"^(\d{8})_Taux[ _]?No[eé]mie.*\.xlsx$", re.IGNORECASE)
RECAP_SHEET        = "Taux_Noémisation"

# Colonnes attendues dans l'extraction, repérées par libellé normalisé et
# jamais par lettre : la position dérive dès qu'une colonne est ajoutée en amont.
COL_LIBCRT = "libcrt"
COL_ACTIFS = "actifs"
COL_NONACT = "non actifs"
COL_NONNOE = "non noemisable"

# Bloc F→L : présent seulement si Laurence a déjà complété le fichier.
COL_ACTIFS2 = "actifs2"
COL_NONACT2 = "non actifs2"
COL_NONNOE2 = "non noemisable2"
COL_TAUX    = "taux"
COL_OFFRES  = "offres"
COL_INDCOL  = "individuel ou collectif"
COL_ORDRE   = "ordre"

# Libellés d'en-tête du bloc, repris à l'identique du fichier de référence
# (casse comprise : "Non_Noémisable2" prend bien deux majuscules chez elle).
ENTETES_BLOC = ["Actifs2", "Non actifs2", "Non_Noémisable2", "Taux", "Offres",
                "Individuel ou Collectif", "ordre"]

# Coquilles connues du libellé 'Offres' : il est saisi à la main et sert
# d'étiquette, pas de clé. Ici "EFS SANTE" désigne le libcrt "EFF SANTE".
# Tout membre non résolu est signalé en [WARN] — ajouter l'alias ici.
ALIAS_MEMBRES: Dict[str, str] = {
    "EFS SANTE":      "EFF SANTE",
    "EFS SANTE BRED": "EFF SANTE BRED",
}

# ─── CONFIG REGROUPEMENTS CLIENT/OFFRE ────────────────────────────────────────
# Figée d'après le fichier de référence 27072026_Taux_Noemie.xlsx complété par
# PILLON Laurence. L'extraction mensuelle ne contient plus ces informations.
#
# Format : (ordre d'affichage, "C" | "I", libellé affiché, [libcrt membres])
#
# Le libellé est PUREMENT DÉCORATIF : les sommes s'appuient sur la liste de
# membres, pas sur un découpage du libellé. C'est volontaire — le libellé de
# référence contient des coquilles (cf. ordre 10 : "EFS SANTE" alors que le
# libcrt de l'extraction est "EFF SANTE"), qui casseraient un appariement
# par nom.
#
# À METTRE À JOUR à chaque nouvelle offre : tout libcrt absent de cette table
# est exclu du récapitulatif ET des totaux, avec une alerte explicite.
# L'ordre du tableau est celui de cette liste (= colonne L du fichier de
# référence). Les deux dernières lignes ont un ordre vide chez Laurence : on
# reproduit la cellule vide, la position en fin de tableau suffit.
NOEMIE_GROUPES: List[Tuple[Optional[int], str, str, List[str]]] = [
    (1,  "C", "NUANCE",                            ["NUANCE"]),
    (2,  "C", "PSC SANTE MEAE",                    ["PSC SANTE MEAE"]),
    (3,  "C", "Education Jeunesse Sports Enseignement Supérieur Recherche",
              ["Education Jeunesse Sports Enseignement Supérieur Recherche"]),
    (4,  "C", "La juridiction administrative",     ["La juridiction administrative"]),
    (5,  "C", "PSC SANTE MINISTERE DE LA CULTURE", ["PSC SANTE MINISTERE DE LA CULTURE"]),
    (6,  "I", "MAEE PR+MAEE SANTE RETRAITE + MAEE PNR",
              ["MAEE PR", "MAEE SANTE RETRAITE", "MAEE PNR"]),
    (7,  "I", "MSP RETRAITES+MGEN SANTE",          ["MSP RETRAITES", "MGEN SANTE"]),
    (8,  "I", "MGEN ALTERNATIVE",                  ["MGEN ALTERNATIVE"]),
    (9,  "I", "MTE ACTIF + MTE",                   ["MTE ACTIF", "MTE"]),
    # Libellé "EFS" conservé tel quel (affichage attendu par la direction) ;
    # les libcrt réels de l'extraction sont bien "EFF SANTE" / "EFF SANTE BRED".
    (10, "I", "EFS SANTE + EFS SANTE BRED",        ["EFF SANTE", "EFF SANTE BRED"]),
    (11, "I", "MGEN OJI",                          ["MGEN OJI"]),
    (12, "I", "MISP ACTIF + MISP + MISP R",        ["MISP ACTIF", "MISP", "MISP R"]),
    (13, "I", "C2S PF + C2S",                      ["C2S PF", "C2S"]),
    (14, "I", "C2S SORTIE",                        ["C2S SORTIE"]),
    # Rattachement "I" repris du fichier de référence 27072026_Taux_Noemie 2.
    # ATTENTION : une autre version du même fichier les classe en "C", ce qui
    # fait passer le taux Collectif de 83,63 % à 83,61 %. À confirmer avec
    # Laurence. Sans effet tant qu'elle fournit le bloc F→L : dans ce cas le
    # rattachement est lu dans le fichier, cette table n'étant qu'un repli.
    (None, "I", "PSC SANTE MIOM",                  ["PSC SANTE MIOM"]),
    (None, "I", "PSC SANTE MSST",                  ["PSC SANTE MSST"]),
]

# ─── FORMATS / STYLES (alignés sur 05_generation_tcd.py) ──────────────────────

FMT_INT = "#,##0"
FMT_PCT = "0.00%"

_HEADER_FILL = PatternFill("solid", fgColor="1F497D")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_GROUP_FILL  = PatternFill("solid", fgColor="E2EFDA")   # vert du bloc F→L
_TOTAL_FILL  = PatternFill("solid", fgColor="DCE6F1")
_TOTAL_FONT  = Font(bold=True, size=10)
_GRAND_FILL  = PatternFill("solid", fgColor="FFF2CC")
_GRAND_FONT  = Font(bold=True, size=10, color="7F6000")
_DATA_FONT   = Font(size=10)
_CENTER      = Alignment(horizontal="center", vertical="center")
_LEFT        = Alignment(horizontal="left",   vertical="center")
_TAB_COLOR   = "70AD47"

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def _norm(value) -> str:
    """Normalise un libellé pour comparaison : sans accent, minuscule, espaces
    simples. '_' est traité comme un espace ('Non_noémisable' → 'non noemisable')."""
    if value is None:
        return ""
    txt = unicodedata.normalize("NFKD", str(value))
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    txt = txt.replace("_", " ").replace("\xa0", " ").lower()
    return re.sub(r"\s+", " ", txt).strip()


def _num(value) -> float:
    """Convertit une cellule Excel en float ; cellule vide/texte → 0.0."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _taux(actifs: float, non_actifs: float, non_noem: float) -> Optional[float]:
    """Taux de noémisation. None si le dénominateur est nul (offre sans contrat)."""
    denom = actifs + non_actifs + non_noem
    return actifs / denom if denom else None


def _fmt_int(value: float) -> str:
    """Entier avec séparateur d'espace, pour les traces console."""
    return f"{int(value):,}".replace(",", " ")


def _split_membres(label: str) -> List[str]:
    """Éclate un libellé de regroupement ('A + B+C') en offres membres."""
    return [m.strip() for m in str(label).split("+") if m.strip()]

# ─── LECTURE DE L'EXTRACTION ──────────────────────────────────────────────────

def find_input_file() -> Optional[Path]:
    """Localise l'extraction Taux_Noémie la plus récente. None si absente
    (cas fonctionnel accepté : l'étape est simplement ignorée)."""
    matches = [f for f in INPUT_DIR.iterdir()
               if f.is_file() and not f.name.startswith("~$")
               and NOEMIE_FILENAME_RE.match(f.name)]
    if not matches:
        return None
    if len(matches) > 1:
        matches.sort(key=lambda f: NOEMIE_FILENAME_RE.match(f.name).group(1), reverse=True)
        warnings.warn(
            f"[WARN] Plusieurs extractions Taux_Noémie détectées : "
            f"{[f.name for f in matches]}\n  → Utilisation de {matches[0].name}",
            stacklevel=2,
        )
    return matches[0]


def find_header_row(ws) -> int:
    """Ligne d'en-tête = 1ʳᵉ ligne contenant 'libcrt'.

    L'extraction Power BI place un rappel des filtres en ligne 1 et une ligne
    vide en ligne 2 ; l'en-tête est donc en ligne 3 — mais rien ne garantit
    que ce cadrage reste stable d'un mois sur l'autre.
    """
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30)):
        for cell in row:
            if _norm(cell.value) == COL_LIBCRT:
                return cell.row
    raise ValueError(
        f"En-tête introuvable dans '{ws.title}' : aucune cellule '{COL_LIBCRT}' "
        f"dans les 30 premières lignes.\n"
        f"  → Vérifier que l'extraction Power BI n'a pas changé de structure."
    )


def map_columns(ws, header_row: int) -> Dict[str, int]:
    """Libellé normalisé → index de colonne (1-based)."""
    mapping: Dict[str, int] = {}
    for cell in ws[header_row]:
        key = _norm(cell.value)
        if key and key not in mapping:
            mapping[key] = cell.column
    return mapping


def read_data_rows(ws, header_row: int, cols: Dict[str, int]) -> Dict[str, dict]:
    """Lit les lignes de données jusqu'aux totaux (exclus) → {libcrt: valeurs}.

    Fin de bloc = 1ʳᵉ ligne dont le libcrt est vide ou commence par 'total'
    (l'extraction peut ou non embarquer ses propres lignes de totaux).
    """
    c_lib = cols[COL_LIBCRT]
    rows: Dict[str, dict] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        libcrt = ws.cell(row=r, column=c_lib).value
        norm = _norm(libcrt)
        if not norm or norm.startswith("total"):
            break
        libelle = str(libcrt).strip()

        def lire(cle):
            """Valeur brute d'une colonne du bloc F→L, None si absente."""
            return ws.cell(row=r, column=cols[cle]).value if cle in cols else None

        if libelle in rows:
            print(f"  [WARN]    libcrt '{libelle}' présent en double "
                  f"(ligne {r}) — valeurs cumulées.")
            rows[libelle]["actifs"]  += _num(ws.cell(row=r, column=cols[COL_ACTIFS]).value)
            rows[libelle]["non_act"] += _num(ws.cell(row=r, column=cols[COL_NONACT]).value)
            rows[libelle]["non_noe"] += _num(ws.cell(row=r, column=cols[COL_NONNOE]).value)
            continue
        rows[libelle] = {
            "actifs":  _num(ws.cell(row=r, column=cols[COL_ACTIFS]).value),
            "non_act": _num(ws.cell(row=r, column=cols[COL_NONACT]).value),
            "non_noe": _num(ws.cell(row=r, column=cols[COL_NONNOE]).value),
            # Bloc F→L éventuel — laissé brut (None = cellule vide).
            "actifs2": lire(COL_ACTIFS2), "non_act2": lire(COL_NONACT2),
            "non_noe2": lire(COL_NONNOE2), "taux": lire(COL_TAUX),
            "offres": lire(COL_OFFRES), "indcol": lire(COL_INDCOL),
            "ordre": lire(COL_ORDRE),
        }
    return rows

# ─── CONSOLIDATION ────────────────────────────────────────────────────────────

def controler_couverture(rows: Dict[str, dict]) -> List[str]:
    """Confronte les libcrt de l'extraction à la config.

    Deux écarts possibles, tous deux non bloquants mais signalés :
      - libcrt inconnu   → offre nouvelle, exclue du récap ET des totaux ;
      - membre absent    → offre disparue du flux, somme du groupe réduite.
    Retourne la liste des libcrt inconnus.
    """
    connus = {m for _o, _ic, _lb, membres in NOEMIE_GROUPES for m in membres}

    doublons = [m for m in connus
                if sum(m in membres for *_x, membres in NOEMIE_GROUPES) > 1]
    if doublons:
        print(f"  [WARN]    Config incohérente : {sorted(set(doublons))} "
              f"rattaché(s) à plusieurs groupes — double comptage.")

    absents = sorted(connus - set(rows))
    if absents:
        print(f"  [WARN]    {len(absents)} offre(s) de la config absente(s) du "
              f"flux : {absents}\n"
              f"            → Somme du regroupement réduite d'autant.")

    inconnus = sorted(set(rows) - connus)
    if inconnus:
        perdus = sum(rows[k]["actifs"] + rows[k]["non_act"] + rows[k]["non_noe"]
                     for k in inconnus)
        print("\n" + "!" * 60)
        print(f"  [ALERTE]  {len(inconnus)} libcrt absent(s) de NOEMIE_GROUPES — "
              f"EXCLU(S) du récapitulatif et des totaux :")
        for k in inconnus:
            v = rows[k]
            print(f"            {k:<40} actifs={_fmt_int(v['actifs'])} "
                  f"non actifs={_fmt_int(v['non_act'])} "
                  f"non noémisables={_fmt_int(v['non_noe'])}")
        print(f"            Soit {_fmt_int(perdus)} contrats non comptabilisés.")
        print("            → Ajouter ces offres à NOEMIE_GROUPES "
              "(demander à Laurence leur rattachement C/I et leur ordre).")
        print("!" * 60 + "\n")
    return inconnus


def _index_config() -> Dict[str, Tuple[Optional[int], str, str, str]]:
    """libcrt (n'importe quel membre) → (ordre, C/I, libellé, porteuse) config."""
    idx: Dict[str, Tuple[Optional[int], str, str, str]] = {}
    for ordre, ic, libelle, membres in NOEMIE_GROUPES:
        for m in membres:
            idx[m] = (ordre, ic, libelle, membres[0])
    return idx


def fichier_definit_groupes(rows: Dict[str, dict]) -> bool:
    """La colonne 'Offres' du fichier est-elle renseignée ?

    Seul le libellé compte : les colonnes Actifs2/Non actifs2/Non_Noémisable2
    du fichier ne sont JAMAIS relues. Elles portent des formules Excel dont le
    résultat en cache peut être périmé — modifier le libellé d'un regroupement
    ne met pas le cache à jour tant qu'Excel n'a pas recalculé. Les sommes sont
    donc systématiquement recalculées depuis les colonnes A→E.
    """
    return any(r.get("offres") for r in rows.values())


def _membres_du_libelle(libcrt: str, libelle, index_norm: Dict[str, str]
                        ) -> Tuple[List[str], List[str]]:
    """Libellé 'Offres' → (libcrt membres résolus, jetons non résolus).

    Le libellé est saisi à la main : on normalise (accents, casse, espaces) et
    on applique ALIAS_MEMBRES avant d'abandonner un jeton. La porteuse est
    toujours incluse, même absente de son propre libellé.
    """
    membres, orphelins = [], []
    for jeton in _split_membres(libelle) if libelle else []:
        cible = ALIAS_MEMBRES.get(jeton.upper(), jeton)
        resolu = index_norm.get(_norm(cible))
        (membres if resolu else orphelins).append(resolu or jeton)
    if libcrt not in membres:
        membres.insert(0, libcrt)
    return list(dict.fromkeys(membres)), orphelins


def _verifier_partition(groupes: List[dict], rows: Dict[str, dict]) -> None:
    """Chaque libcrt doit relever d'exactement un groupe.

    Un libcrt cité par deux regroupements est compté deux fois dans les
    totaux ; un libcrt cité par aucun disparaît du tableau. Les deux cas sont
    silencieux à la lecture — d'où l'alerte.
    """
    proprietaires: Dict[str, List[str]] = {}
    for g in groupes:
        for m in g.get("membres", []):
            proprietaires.setdefault(m, []).append(g["libelle"])

    for libcrt, sources in proprietaires.items():
        if len(sources) > 1:
            print(f"  [ALERTE]  '{libcrt}' rattaché à {len(sources)} "
                  f"regroupements : {sources}\n"
                  f"            → Compté {len(sources)} fois dans les totaux. "
                  f"Corriger la colonne 'Offres'.")

    absents = [k for k in rows if k not in proprietaires]
    if absents:
        perdus = sum(rows[k]["actifs"] + rows[k]["non_act"] + rows[k]["non_noe"]
                     for k in absents)
        print(f"  [ALERTE]  {len(absents)} libcrt cité(s) par aucun "
              f"regroupement : {absents}\n"
              f"            → {_fmt_int(perdus)} contrats exclus du tableau "
              f"et des totaux.")


def consolider(rows: Dict[str, dict]) -> Tuple[List[dict], Dict[str, str]]:
    """Consolide par client/offre, colonne par colonne.

    Chacune des trois informations est prise dans le fichier si Laurence l'a
    fournie, sinon dans NOEMIE_GROUPES. Les trois sont indépendantes : elle
    peut envoyer les offres sans l'ordre, le rattachement C/I seul, etc.

      Offres                   → regroupement des lignes
      Individuel ou Collectif  → ventilation des totaux C / I
      ordre                    → ordre d'affichage du tableau

    Retourne (groupes, origine de chaque colonne) pour tracer la décision.
    """
    idx = _index_config()
    groupes: List[dict] = []
    origine = {"offres": "config", "indcol": "config", "ordre": "config"}

    if fichier_definit_groupes(rows):
        origine["offres"] = "fichier"
        index_norm = {_norm(k): k for k in rows}
        # Ligne porteuse = libellé 'Offres' OU 'ordre' renseigné. Les deux
        # critères sont nécessaires : dans le fichier de référence la 1ʳᵉ ligne
        # (NUANCE) a un ordre sans libellé, et les deux dernières
        # (PSC SANTE MIOM/MSST) un libellé sans ordre.
        for libcrt, r in rows.items():
            if not (r.get("offres") or r.get("ordre") is not None):
                continue                  # ligne membre : agrégée chez sa porteuse
            membres, orphelins = _membres_du_libelle(libcrt, r.get("offres"),
                                                     index_norm)
            if orphelins:
                print(f"  [WARN]    Regroupement '{r['offres']}' : offre(s) "
                      f"inconnue(s) {orphelins} — exclue(s) de la somme.\n"
                      f"            → Vérifier l'orthographe, ou ajouter "
                      f"l'alias dans ALIAS_MEMBRES.")
            # Sommes recalculées depuis A→E : les colonnes Actifs2/… du fichier
            # sont ignorées, leur cache Excel pouvant être périmé.
            actifs  = sum(rows[m]["actifs"]  for m in membres)
            non_act = sum(rows[m]["non_act"] for m in membres)
            non_noe = sum(rows[m]["non_noe"] for m in membres)
            groupes.append({
                "porteur": libcrt, "membres": membres,
                "regroupement": len(membres) > 1,
                "libelle": str(r["offres"]).strip() if r.get("offres")
                           else idx.get(libcrt, (None, "", libcrt, ""))[2],
                "actifs": actifs, "non_act": non_act, "non_noe": non_noe,
                "taux": _taux(actifs, non_act, non_noe),
            })
        _verifier_partition(groupes, rows)
    else:
        controler_couverture(rows)
        for ordre, ic, libelle, membres in NOEMIE_GROUPES:
            presents = [m for m in membres if m in rows]
            actifs  = sum(rows[m]["actifs"]  for m in presents)
            non_act = sum(rows[m]["non_act"] for m in presents)
            non_noe = sum(rows[m]["non_noe"] for m in presents)
            groupes.append({
                "porteur": membres[0], "membres": presents,
                "regroupement": len(presents) > 1, "libelle": libelle,
                "actifs": actifs, "non_act": non_act, "non_noe": non_noe,
                "taux": _taux(actifs, non_act, non_noe),
            })

    # Rattachement C/I et ordre : fichier prioritaire, config en repli, pour
    # chaque groupe indépendamment.
    for g in groupes:
        r   = rows.get(g["porteur"], {})
        cfg = idx.get(g["porteur"], (None, "", "", ""))

        ic_fichier = r.get("indcol")
        if ic_fichier:
            g["indcol"] = str(ic_fichier).strip().upper()[:1]
            origine["indcol"] = "fichier"
        else:
            g["indcol"] = cfg[1]

        ordre_fichier = r.get("ordre")
        if isinstance(ordre_fichier, (int, float)):
            g["ordre"] = int(ordre_fichier)
            origine["ordre"] = "fichier"
        else:
            g["ordre"] = cfg[0]

    # Tri stable : les groupes sans ordre restent en fin, dans leur ordre
    # d'apparition (fichier) ou de déclaration (config).
    groupes.sort(key=lambda g: (g["ordre"] is None, g["ordre"] or 0))
    return groupes, origine


def totaux(groupes: List[dict]) -> Dict[str, dict]:
    """Totaux général / Collectif / Individuel.

    Sommés sur les groupes : la config garantit qu'un libcrt appartient à un
    seul groupe (contrôlé par controler_couverture), donc pas de double
    comptage et Total == Collectif + Individuel par construction.
    """
    def bloc(sel: List[dict]) -> dict:
        a  = sum(g["actifs"]  for g in sel)
        na = sum(g["non_act"] for g in sel)
        nn = sum(g["non_noe"] for g in sel)
        return {"actifs": a, "non_act": na, "non_noe": nn, "taux": _taux(a, na, nn)}

    return {
        "Total":            bloc(groupes),
        "Total Collectif":  bloc([g for g in groupes if g["indcol"] == "C"]),
        "Total Individuel": bloc([g for g in groupes if g["indcol"] == "I"]),
    }

# ─── ÉCRITURE DE LA FEUILLE ───────────────────────────────────────────────────

def write_recap(wb, groupes: List[dict], tot: Dict[str, dict]) -> None:
    """Feuille 'Taux_Noémisation' — reprise à l'identique du bloc vert F→L.

    Colonnes, libellés et mise en forme repris du fichier de référence :
      F Actifs2 · G Non actifs2 · H Non_noémisable2 · I Taux · J Offres
      K Individuel ou Collectif · L ordre

    Convention de Laurence respectée : F/G/H ne sont renseignées que sur les
    regroupements multi-offres. Une offre isolée n'a pas de sous-total à
    afficher — seul son taux figure en colonne I.
    """
    if RECAP_SHEET in wb.sheetnames:
        del wb[RECAP_SHEET]
    ws = wb.create_sheet(RECAP_SHEET)
    ws.sheet_properties.tabColor = _TAB_COLOR

    entetes = ENTETES_BLOC
    for i, libelle in enumerate(entetes, start=1):
        cell = ws.cell(row=1, column=i, value=libelle)
        cell.fill, cell.font, cell.alignment = _HEADER_FILL, _HEADER_FONT, _CENTER

    for i, g in enumerate(groupes, start=2):
        regroupement = g["regroupement"]
        valeurs = [
            (1, g["actifs"]  if regroupement else None, FMT_INT),
            (2, g["non_act"] if regroupement else None, FMT_INT),
            (3, g["non_noe"] if regroupement else None, FMT_INT),
            (4, g["taux"],    FMT_PCT),
            (5, g["libelle"], None),
            (6, g["indcol"],  None),
            (7, g["ordre"],   None),
        ]
        for col, value, fmt in valeurs:
            cell = ws.cell(row=i, column=col, value=value)
            cell.font = _DATA_FONT
            cell.fill = _GROUP_FILL
            cell.alignment = _LEFT if col == 5 else _CENTER
            if fmt:
                cell.number_format = fmt

    # Totaux : les 3 indicateurs remontés au dashboard, alignés sur les
    # colonnes agrégées du bloc (F/G/H/I).
    start = len(groupes) + 3
    for i, libelle in enumerate(["Total", "Total Collectif", "Total Individuel"]):
        r, t  = start + i, tot[libelle]
        grand = (libelle == "Total")
        fill  = _GRAND_FILL if grand else _TOTAL_FILL
        font  = _GRAND_FONT if grand else _TOTAL_FONT
        valeurs = [(1, t["actifs"], FMT_INT), (2, t["non_act"], FMT_INT),
                   (3, t["non_noe"], FMT_INT), (4, t["taux"], FMT_PCT),
                   (5, libelle, None)]
        for col, value, fmt in valeurs:
            cell = ws.cell(row=r, column=col, value=value)
            cell.font, cell.fill = font, fill
            cell.alignment = _LEFT if col == 5 else _CENTER
            if fmt:
                cell.number_format = fmt

    ws.freeze_panes = "A2"
    for i in range(1, len(entetes) + 1):
        largeur = max((len(str(ws.cell(row=r, column=i).value or ""))
                       for r in range(1, ws.max_row + 1)), default=10)
        ws.column_dimensions[get_column_letter(i)].width = min(max(largeur + 2, 12), 60)
    print(f"  [RECAP]   {len(groupes)} clients/offres consolidés")

# ─── MAIN ─────────────────────────────────────────────────────────────────────

def process(input_file: Path) -> Path:
    prefix      = NOEMIE_FILENAME_RE.match(input_file.name).group(1)
    output_file = OUTPUT_DIR / input_file.name

    print(f"\n[INPUT]   {input_file.name}  (préfixe: {prefix})")
    print(f"[OUTPUT]  {output_file}")

    shutil.copy2(input_file, output_file)
    print("\n[COPY]    Fichier dupliqué vers Output/")

    # Deux ouvertures : data_only=True rend les résultats des formules F→I
    # (openpyxl ne calcule rien), l'autre conserve ces formules dans la copie
    # de sortie — sauvegarder le classeur data_only les remplacerait par des
    # valeurs figées et détruirait le fichier de Laurence.
    wb_val = load_workbook(output_file, data_only=True)
    ws_val = wb_val[wb_val.sheetnames[0]]
    header_row = find_header_row(ws_val)
    cols = map_columns(ws_val, header_row)
    print(f"[LOAD]    Feuille '{ws_val.title}' — en-tête ligne {header_row}")

    manquantes = [c for c in (COL_LIBCRT, COL_ACTIFS, COL_NONACT, COL_NONNOE)
                  if c not in cols]
    if manquantes:
        raise ValueError(
            f"Colonnes obligatoires absentes de l'extraction : {manquantes}\n"
            f"  Colonnes trouvées : {sorted(cols)}"
        )

    rows = read_data_rows(ws_val, header_row, cols)
    wb_val.close()
    print(f"[PARSER]  {len(rows)} libcrt lus")

    print("\n[CALCUL]  Consolidation par client/offre...")
    groupes, origine = consolider(rows)
    print(f"  [SOURCE]  offres={origine['offres']}  "
          f"individuel/collectif={origine['indcol']}  ordre={origine['ordre']}"
          f"   (fichier = fourni par Laurence, config = valeur par défaut)")
    tot = totaux(groupes)

    wb = load_workbook(output_file)

    print("\n[FEUILLE] Génération du récapitulatif...")
    write_recap(wb, groupes, tot)

    wb.save(output_file)
    wb.close()

    print("\n[KPI]     Indicateurs dashboard :")
    for libelle in ("Total", "Total Collectif", "Total Individuel"):
        t = tot[libelle]
        pct = f"{t['taux']:.2%}" if t["taux"] is not None else "n/a"
        print(f"  {libelle:<18} {pct:>8}   "
              f"(actifs {_fmt_int(t['actifs'])} / "
              f"non actifs {_fmt_int(t['non_act'])} / "
              f"non noémisables {_fmt_int(t['non_noe'])})")

    print(f"\n[DONE]    {output_file.name} sauvegardé "
          f"(feuille '{RECAP_SHEET}' ajoutée).")
    return output_file


def main() -> None:
    print("=" * 60)
    print("09_taux_noemisation.py — Taux de noémisation par client/offre")
    print("=" * 60)

    input_file = find_input_file()
    if input_file is None:
        print(f"\n[INFO]    Aucun fichier '*_Taux_Noémie*.xlsx' trouvé dans {INPUT_DIR}.")
        print("          → Étape taux de noémisation ignorée.")
        print("=" * 60)
        return

    try:
        process(input_file)
    except Exception as exc:
        print(f"\n[ERREUR]  {exc}", file=sys.stderr)
        print("=" * 60)
        sys.exit(1)
    print("=" * 60)


if __name__ == "__main__":
    main()
