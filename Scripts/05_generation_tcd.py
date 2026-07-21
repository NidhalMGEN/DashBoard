# -*- coding: utf-8 -*-
"""
05_generation_tcd.py — Génération automatique des TCD Accolade KPI
===================================================================

Entrée  : Input_Data/{PREFIX}_Accolade - KPI*.xlsx   (PREFIX = DDMMYYYY)
Sortie  : Output/{PREFIX}_Accolade - KPI*.xlsx
          (copie du fichier source avec Synthèse + Feuil1–6 régénérées)

Feuilles générées :
  Synthèse — indicateurs clés agrégés (portefeuille, résiliations, top offres)
  Feuil1   — Résiliations : Société × motif × nombre      (table plate)
  Feuil2   — Résiliations filtrées société                 (TCD par société)
  Feuil3   — offre × codsoc × type_assure × nb_pp         (table plate)
  Feuil4   — nb_pp agrégé par codsoc × type_assure        (TCD avec sous-totaux)
  Feuil7   — listing à plat des offres ASSURE             (regroupements + ordre)
  Feuil5   — offre × codsoc × tranche_âge × nb_pp         (table plate)
  Feuil6   — nb_pp agrégé par codsoc × tranche_âge        (TCD avec sous-totaux)

Usage :
  python 05_generation_tcd.py
"""

from __future__ import annotations

import re
import shutil
import sys
import warnings
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ─── CONFIG ───────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR   = SCRIPT_DIR if (SCRIPT_DIR / "Input_Data").exists() else SCRIPT_DIR.parent
INPUT_DIR  = BASE_DIR / "Input_Data"
OUTPUT_DIR = BASE_DIR / "Output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
INPUT_DIR.mkdir(parents=True, exist_ok=True)

# Fichier "Accolade - KPI*.xlsx" (V1.12, V1.13, 1.13… ) — flux principal TCD.
# Convention de nommage souple : le nom doit contenir "Accolade - KPI" (le
# suffixe de version est libre), à l'exclusion du fichier doublons "Accolade -
# KPI 4*" qui dispose de son propre flux (cf. KPI4_FILENAME_RE).
KPI_FILENAME_RE = re.compile(r"^(\d{8})_(Accolade - KPI(?! 4(?!\.\d)).*\.xlsx)$", re.IGNORECASE)
# La feuille principale est repérée par le même motif souple "Accolade - KPI"
# (hors "Accolade - KPI 4") plutôt que par un nom de version figé.
# L'exclusion (?! 4(?!\.\d)) écarte le flux doublons "KPI 4" tout en conservant
# une éventuelle version majeure 4 du flux principal (ex. "KPI 4.0").
MAIN_SHEET_RE   = re.compile(r"^Accolade - KPI(?! 4(?!\.\d))", re.IGNORECASE)
TCD_SHEETS      = ["Synthèse", "Feuil1", "Feuil2", "Feuil3", "Feuil4", "Feuil7", "Feuil5", "Feuil6"]

# Fichier "Accolade - KPI 4*.xlsx" (V2.0+) — KPI doublons de contrats
KPI4_FILENAME_RE      = re.compile(r"^(\d{8})_(Accolade - KPI 4.*\.xlsx)$", re.IGNORECASE)
KPI4_MAIN_SHEET_RE    = re.compile(r"^Accolade - KPI 4", re.IGNORECASE)
KPI4_OUTPUT_SHEET     = "KPI_doublon_contrat"
KPI4_DATA_HEADER_ROW  = 18
KPI4_HEADER_SCAN_ROWS = 17

# Marqueurs de section : valeur exacte de la cellule d'en-tête du bloc (insensible à la casse)
MARKER_RESILIATIONS = "motif resiliation"
MARKER_TYPE_ASSURE  = "type assure"
MARKER_TRANCHE_AGE  = "tranche age"

# Colonnes numériques attendues par bloc (conversion explicite)
NUM_COLS_RESIL    = {"nombre de résiliations"}
NUM_COLS_TYPE_AGE = {"nb pp a date", "nb pp non radie"}

# Ordre DWH des codes de résiliation — imposé dans le TCD Feuil2 et la Synthèse.
# Les codes hors liste sont silencieusement ignorés (cf. décision projet).
RESILIATION_DWH_ORDER = [
    "AGE", "ANNUL", "CCULT", "CJUAD", "CMEAE", "CMEN", "CMGAS", "COLLO",
    "CONC", "COT+", "CSS", "DECES", "DEM", "DEMUT", "DENT", "DETRA", "DISP",
    "DOUBL", "DRET", "FMGAR", "FPORT", "IMP", "OFFRE", "PREST", "RADIA",
    "SUSP", "TRANS",
]

# Libellé des résiliations sans motif renseigné — doit être comptabilisé dans
# le total général (cohérence avec le TCD source Excel qui affiche "(vide)").
RESIL_EMPTY_LABEL = "(vide)"
RESILIATION_ORDER_WITH_EMPTY = RESILIATION_DWH_ORDER + [RESIL_EMPTY_LABEL]

# Schémas minimaux pour validation
SCHEMA = {
    "résiliations": {"société", "motif resiliation", "nombre de résiliations"},
    "type_assure":  {"offre", "codsoc", "type assure", "nb pp a date", "nb pp non radie"},
    "tranche_âge":  {"offre", "codsoc", "tranche age", "nb pp a date", "nb pp non radie"},
}

# ─── CONFIG GROUPEMENTS OFFRES (Feuil4 bloc droit) ────────────────────────────
# Format : "OFFRE" → (numéro_groupe | None, libellé_groupe | None)
# À METTRE À JOUR si de nouvelles offres apparaissent dans un flux futur.
OFFRE_GROUPES: Dict[str, tuple] = {
    # Regroupement INMMISP* — total porté par la 1ʳᵉ offre (INMMISPACT).
    "INMMISPACT":  (28,   "INMMISPACT + INMMISPAUT + INMMISPRET"),
    "INMMSPACT1":  (32,   None),
    "INMMSPOJI1":  (29,   None),
    "INMMSPRET1":  (1,    None),
    "INPCULT002":  (3,    "INPCULT002 + MEPCULT003"),
    "INPPREVIND":  (25,   None),
    # Regroupement INSC2S* — toutes les offres comptent (aucune exclusion).
    "INSC2SG":     (30,   "INSC2SG + INSC2SP + INSC2SS"),
    "INSEFFBRED":  (2,    "INSEFFBRED + INSEFFICIE"),
    "INSMAEEREF":  (4,    "INSMAEEPNR+INSMAEEREF + INSMAEERET"),
    # Demande PILLON Laurence (30/06) : INSMALT001 en ordre 33, total en 34.
    "INSMALT001":  (33,   None),
    "INSMAS001":   (22,   None),
    # Regroupement INSMTE* — total porté par INSMTE0001.
    "INSMTE0001":  (31,   "INSMTE0001 + INSMTE1001"),
    "INSMIOM001":  (24,   None),
    "MEPMAS001":   (21,   None),
    "MEPMEAE001":  (23,   "MEPMEAE001+MEPMEAT002"),
    "MEPMENP002":  (27,   "MEPMENP002+MEPMENP201"),
    "MESAPEI002":  (5,    None),
    "MESBGTL004":  (6,    None),
    "MESBSRT002":  (7,    None),
    "MESCULT002":  (8,    "MESCULT002 +MESCULT101"),
    "MESERASMUS":  (9,    None),
    "MESESEA001":  (10,   None),
    "MESFRVL001":  (11,   None),
    "MESGMMS004":  (12,   None),
    "MESINOV003":  (13,   None),
    "MESJUAD002":  (14,   "MESJUAD002+MESJUAM005"),
    "MESLAD46":    (15,   None),
    "MESMEAE102":  (16,   "MESMEAE102 + MESMEAE202 + MESMEAE301"),
    "MESMEN001":   (17,   "MESMEN001 + MESMEN2003 + MESMEN3001"),
    "MESMNTSALA":  (18,   None),
    "MESMUTFR31":  (19,   None),
    "OFFSUP":      (None, None),   # offre standalone (2 pp)
    "STSNUANCE":   (20,   None),
}


def _groupes_effectifs(offres_presentes: set) -> Dict[str, tuple]:
    """OFFRE_GROUPES recalculé sur les offres réellement présentes dans le flux.

    Si l'offre clé d'un regroupement a disparu du flux (ex: INSMAEEREF tombée
    à 0 assuré au 30/06), le total, le libellé et l'ordre d'affichage sont
    portés par le premier membre du groupe encore présent — sinon le
    regroupement disparaîtrait silencieusement des synthèses Feuil4/Feuil7.
    """
    eff: Dict[str, tuple] = {}
    for key, (num, label) in OFFRE_GROUPES.items():
        carrier = key
        if key not in offres_presentes:
            if not label:
                continue  # offre standalone absente du flux
            presents = [m for m in (x.strip() for x in label.split("+"))
                        if m in offres_presentes]
            if not presents:
                continue  # groupe entièrement absent du flux
            carrier = presents[0]
            print(f"  [WARN] Offre clé '{key}' absente du flux → regroupement "
                  f"'{label}' porté par '{carrier}'")
        eff[carrier] = (num, label)
    return eff


def _tri_offre_key(offre, groupes: Dict[str, tuple]):
    """Clé de tri des offres : ordre croissant ; offres sans ordre (clé vide)
    reléguées en fin de bloc (départage alphabétique pour la reproductibilité)."""
    ordre = groupes.get(offre, (None, None))[0]
    return (ordre is None, ordre or 0, str(offre))

# ─── FORMATS EXCEL ────────────────────────────────────────────────────────────
FMT_INT = "#,##0"
FMT_PCT = "0.00%"

# ─── STYLES ───────────────────────────────────────────────────────────────────

_HEADER_FILL      = PatternFill("solid", fgColor="1F497D")
_HEADER_FONT      = Font(bold=True, color="FFFFFF", size=10)
_TOTAL_FILL       = PatternFill("solid", fgColor="DCE6F1")
_TOTAL_FONT       = Font(bold=True, size=10)
_GRAND_TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
_GRAND_TOTAL_FONT = Font(bold=True, size=10, color="7F6000")
_DATA_FONT        = Font(size=10)
_PCT_FONT         = Font(size=10, italic=True, color="595959")
_CENTER           = Alignment(horizontal="center", vertical="center", wrap_text=False)
_LEFT             = Alignment(horizontal="left",   vertical="center", wrap_text=False)

# Couleurs des onglets
_TAB_FLAT  = "4472C4"   # bleu  — tables plates (Feuil1, 3, 5)
_TAB_TCD   = "ED7D31"   # orange — TCD (Feuil2, 4, 6)
_TAB_SYNTH = "70AD47"   # vert  — Synthèse

# Filtre codsoc standards (exclut les codes non-standard comme '000')
VALID_CODSOC = re.compile(r'^0[1-9]\d$')

# ─── KPI HELPERS ──────────────────────────────────────────────────────────────

def _pct(value, total) -> Optional[float]:
    """Retourne value/total (0..1) compatible avec number_format='0.0%', ou None."""
    try:
        return float(value) / float(total) if total else None
    except (TypeError, ZeroDivisionError):
        return None


def _taux_rad(nb_adate, nb_nonrad) -> Optional[float]:
    """
    Taux de radiation = (nb_non_radié − nb_à_date) / nb_non_radié
    Interprétation : % des contrats historiques qui ne sont plus actifs aujourd'hui.
    """
    try:
        return (float(nb_nonrad) - float(nb_adate)) / float(nb_nonrad) if nb_nonrad else None
    except (TypeError, ZeroDivisionError):
        return None

# ─── UTILITAIRES I/O ──────────────────────────────────────────────────────────

def find_input_file() -> Optional[Path]:
    """Recherche le fichier Accolade - KPI dans INPUT_DIR.

    Retourne None si aucun fichier n'est trouvé (cas fonctionnel accepté :
    l'absence de fichier dans Input_Data/ est le seul critere de skip,
    sans condition de jour). Le main() gérera l'absence proprement.
    """
    matches = [f for f in INPUT_DIR.iterdir() if KPI_FILENAME_RE.match(f.name)]
    if not matches:
        return None
    if len(matches) > 1:
        # Trier par préfixe de date (DDMMYYYY) décroissant → fichier le plus récent
        matches.sort(key=lambda f: KPI_FILENAME_RE.match(f.name).group(1), reverse=True)
        names = [f.name for f in matches]
        warnings.warn(
            f"[WARN] Plusieurs fichiers KPI détectés : {names}\n"
            f"  → Utilisation du plus récent : {matches[0].name}",
            stacklevel=2,
        )
    return matches[0]


def extract_prefix(filename: str) -> str:
    m = KPI_FILENAME_RE.match(filename)
    if not m:
        raise ValueError(f"Impossible d'extraire le préfixe depuis '{filename}'")
    return m.group(1)


def find_main_sheet(wb) -> str:
    """Retourne le nom de la feuille principale 'Accolade - KPI*' (hors 'KPI 4')."""
    for name in wb.sheetnames:
        if MAIN_SHEET_RE.match(name):
            return name
    raise ValueError(
        f"Feuille principale 'Accolade - KPI*' introuvable.\n"
        f"Feuilles disponibles : {wb.sheetnames}"
    )

# ─── PARSER ───────────────────────────────────────────────────────────────────

# Patterns de recherche des KPI scalaires dans la feuille principale.
# Chaque entrée : regex (appliquée au texte lower) → clé du dict résultat.
#
# On s'ancre sur le NUMÉRO du KPI ("KPI 1", "KPI 2 Bis"…) et non sur le libellé.
# Motif : le libellé est rédactionnel et change côté producteur du flux. Le
# 17/07/2026 (source V1.14) le mot "AI" a été remplacé par "Accolade" dans les
# titres KPI 1, 2 et 2 Bis ; les anciennes sous-chaînes ("nouvelles
# souscriptions dans ai"…) ne matchaient plus, parse_kpi_scalars() renvoyait {}
# et tout le bloc "KPI issus de la feuille source" disparaissait en silence de
# la Synthèse. Le numéro de KPI, lui, est stable : c'est le contrat avec le flux.
#
# Les lookahead (?!\s*(bis|ter)) évitent qu'un motif "KPI 2" capte "KPI 2 Bis"
# ou "KPI 2 Ter" (ce dernier n'est pas repris dans la Synthèse).
_KPI_SCALAR_PATTERNS: List[tuple] = [
    (re.compile(r"^kpi\s*1\b(?!\s*(bis|ter))"), "kpi1_nouvelles_souscriptions"),
    (re.compile(r"^kpi\s*2\s*bis\b"),           "kpi2bis_resiliations_2024"),
    (re.compile(r"^kpi\s*2\b(?!\s*(bis|ter))"), "kpi2_resiliations_total"),
    # KPI 5 / 5 Bis / 6 : volontairement non extraits — ils n'ont jamais figuré
    # dans la Synthèse (leurs anciennes sous-chaînes ne correspondaient déjà pas
    # au libellé source). Décommenter pour les ajouter au bloc ; les libellés
    # correspondants existent déjà dans _KPI_LABELS (cf. write_synthese).
    # (re.compile(r"^kpi\s*5\s*bis\b"),           "kpi5bis_sans_kpep_total"),
    # (re.compile(r"^kpi\s*5\b(?!\s*(bis|ter))"), "kpi5_sans_kpep_2024"),
    # (re.compile(r"^kpi\s*6\b(?!\s*(bis|ter))"), "kpi6_sans_mail"),
]

# KPI scalaires attendus dans tout flux. Leur absence signale un changement de
# structure du fichier source et doit être visible dans le log (cf. incident du
# 17/07/2026 : échec silencieux pendant plusieurs jours).
_KPI_SCALAR_REQUIRED = {key for _rx, key in _KPI_SCALAR_PATTERNS}


def parse_kpi_scalars(ws) -> Dict[str, float]:
    """
    Extrait les valeurs scalaires des KPI depuis la feuille principale
    en cherchant chaque pattern de titre puis la première valeur numérique
    positive sur la même ligne à droite du titre.
    """
    result: Dict[str, float] = {}
    for row in ws.iter_rows():
        for cell in row:
            if not cell.value:
                continue
            cell_text = str(cell.value).strip().lower()
            for pattern, key in _KPI_SCALAR_PATTERNS:
                if key not in result and pattern.search(cell_text):
                    for col in range(cell.column + 1, ws.max_column + 1):
                        v = ws.cell(row=cell.row, column=col).value
                        if isinstance(v, (int, float)) and v > 0:
                            result[key] = float(v)
                            break
    return result


# Patterns KPI 8 : sous-chaîne (lower) → clé.
# Les valeurs sont DANS le texte de la cellule, après ": ".
_KPI8_PATTERNS: List[tuple] = [
    ("déphasages mails",                "dep_mail"),
    ('téléphone "port"',                "dep_tel_port"),
    ('téléphone "telp"',                "dep_tel_telp"),
    ("déphasages dates de naissance",   "dep_naissance"),
    ("déphasages dates de dc",          "dep_deces"),
    ("déphasages d'adresses",           "dep_adresses"),
]


def parse_kpi8_dephasages(ws) -> Dict[str, float]:
    """
    Extrait les compteurs de déphasage KPI 8 depuis le texte des cellules.
    Les valeurs sont intégrées dans la phrase (ex: "... mails : 40047").
    """
    result: Dict[str, float] = {}
    for row in ws.iter_rows():
        for cell in row:
            if not cell.value:
                continue
            text = str(cell.value).strip()
            text_lower = text.lower()
            for pattern, key in _KPI8_PATTERNS:
                if key not in result and pattern in text_lower:
                    m = re.search(r":\s*(\d[\d\s\u202f]*)\s*$", text)
                    if m:
                        val = int(re.sub(r"[\s\u202f]", "", m.group(1)))
                        result[key] = float(val)
    return result


def find_marker_row(ws, marker: str) -> int:
    """
    Retourne le numéro de ligne (1-based) de la première cellule dont la valeur
    correspond exactement à `marker` (insensible à la casse et aux espaces).
    """
    target = marker.strip().lower()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip().lower() == target:
                return cell.row
    raise ValueError(
        f"Marqueur '{marker}' introuvable dans la feuille '{ws.title}'.\n"
        f"Vérifiez que la feuille principale n'a pas changé de structure."
    )


def extract_block(ws, header_row: int, num_cols: set) -> pd.DataFrame:
    """
    Extrait un bloc tabulaire dont l'en-tête est sur `header_row`.
    Collecte TOUTES les cellules non-vides de la ligne d'en-tête (gestion des gaps).
    """
    col_indices: List[int] = []
    headers: List[str] = []
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val is not None:
            col_indices.append(col_idx)
            headers.append(str(val).strip())

    if not headers:
        raise ValueError(f"Aucun en-tête détecté à la ligne {header_row} de '{ws.title}'")

    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_data = [ws.cell(row=row_idx, column=c).value for c in col_indices]
        if all(v is None for v in row_data):
            break
        rows.append(row_data)

    df = pd.DataFrame(rows, columns=headers).dropna(how="all").reset_index(drop=True)
    for col in df.columns:
        if col.lower() in {c.lower() for c in num_cols}:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def _normalize_codsoc(series: pd.Series) -> pd.Series:
    """Normalise codsoc en chaîne zfill(3) (ex: 10 → '010')."""
    return series.astype(str).str.strip().str.zfill(3)


def validate_block(df: pd.DataFrame, block_name: str) -> None:
    expected = SCHEMA[block_name]
    actual   = {c.lower() for c in df.columns}
    missing  = {c.lower() for c in expected} - actual
    if missing:
        raise ValueError(
            f"[VALIDATION] Bloc '{block_name}' : colonnes manquantes : {missing}\n"
            f"  Colonnes détectées : {list(df.columns)}"
        )


def validate_data(df_type: pd.DataFrame) -> None:
    """
    Contrôles de robustesse sur le bloc type_assure :
    - Alerte si des offres inconnues apparaissent (absentes de OFFRE_GROUPES).
    - Affiche le total pp non radie (ASSURE) pour cross-check manuel.
    """
    col_offre = next(c for c in df_type.columns if "offre" in c.lower())
    col_type  = next(c for c in df_type.columns if "type" in c.lower())
    col_nonr  = next(c for c in df_type.columns if "non radie" in c.lower())

    # Références valides = clés + membres de groupes (dans les libellés "+")
    offres_connues = set(OFFRE_GROUPES.keys())
    for _, grp_label in OFFRE_GROUPES.values():
        if grp_label:
            for m in grp_label.split("+"):
                offres_connues.add(m.strip())

    offres_presentes = set(df_type[col_offre].dropna().astype(str).str.strip().unique())
    offres_inconnues = offres_presentes - offres_connues
    if offres_inconnues:
        print(
            f"\n  [WARN] {len(offres_inconnues)} offre(s) absente(s) de OFFRE_GROUPES :\n"
            f"    {sorted(offres_inconnues)}\n"
            f"  → Ajoutez-les à la config OFFRE_GROUPES si elles doivent apparaître "
            f"dans les groupements Feuil4."
        )

    df_assure    = df_type[df_type[col_type].astype(str).str.upper() == "ASSURE"]
    total_assure = df_assure[col_nonr].sum()
    print(f"  [CHECK] Total pp non radie (ASSURE) = {int(total_assure):,}")

# ─── HELPERS ÉCRITURE ─────────────────────────────────────────────────────────

def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0 for c in col),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 55)


def _write_header(ws, row: int, values: list, col_start: int = 1) -> None:
    for offset, val in enumerate(values):
        cell = ws.cell(row=row, column=col_start + offset, value=val)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER


def _write_section_filter(ws, row: int, label: str, value, col_start: int = 1) -> None:
    """Ligne de filtre de section : ex. 'codsoc | 010'."""
    for col_idx, val in enumerate([label, value], start=col_start):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font      = Font(bold=True, size=10)
        cell.alignment = _LEFT


def _write_row(ws, row: int, values: list, total: bool = False, col_start: int = 1,
               num_formats: Optional[Dict[int, str]] = None) -> None:
    """
    Écrit une ligne de données.
    num_formats : dict offset (0-based) → format string Excel,
                  ex: {1: "#,##0", 2: "#,##0", 3: "0.0%"}
    """
    font = _TOTAL_FONT if total else _DATA_FONT
    fill = _TOTAL_FILL if total else None
    for offset, val in enumerate(values):
        col_idx = col_start + offset
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font      = font
        cell.alignment = _LEFT if offset == 0 else _CENTER
        if fill:
            cell.fill = fill
        if num_formats and offset in num_formats:
            cell.number_format = num_formats[offset]

# ─── GÉNÉRATEURS DE FEUILLES ──────────────────────────────────────────────────

def write_feuil1(wb, df: pd.DataFrame) -> None:
    """Table plate résiliations : Société × motif × nombre."""
    ws = wb.create_sheet("Feuil1")
    ws.sheet_properties.tabColor = _TAB_FLAT
    _write_header(ws, 1, list(df.columns))
    num_offsets = {i: FMT_INT for i, c in enumerate(df.columns)
                   if c.lower() in NUM_COLS_RESIL}
    for i, row in df.iterrows():
        _write_row(ws, i + 2, list(row), num_formats=num_offsets)
    ws.freeze_panes = "A2"
    _auto_width(ws)
    print(f"  [FEUIL1] {len(df)} lignes écrites")


def write_feuil2(wb, df: pd.DataFrame) -> None:
    """Résiliations — une section par société + résumé (Toutes)."""
    col_soc   = next(c for c in df.columns if "soci" in c.lower())
    col_motif = next(c for c in df.columns if "motif" in c.lower())
    col_nb    = next(c for c in df.columns if "nombre" in c.lower())

    df = df.copy()
    df[col_nb] = pd.to_numeric(df[col_nb], errors="coerce")
    df[col_motif] = df[col_motif].fillna("").astype(str).str.strip()

    # Les résiliations sans motif renseigné doivent être comptabilisées dans
    # le total général (bucket "(vide)"), comme dans le TCD source.
    df.loc[df[col_motif] == "", col_motif] = RESIL_EMPTY_LABEL

    # Filtrer aux seuls codes DWH + "(vide)" (les autres codes inconnus sont
    # silencieusement ignorés) et imposer l'ordre d'affichage via pd.Categorical.
    df = df[df[col_motif].isin(RESILIATION_ORDER_WITH_EMPTY)].copy()
    df[col_motif] = pd.Categorical(
        df[col_motif], categories=RESILIATION_ORDER_WITH_EMPTY, ordered=True
    )

    HDRS = ["Étiquettes de lignes", "Somme de nombre de résiliations", "% du total"]
    NF   = {1: FMT_INT, 2: FMT_PCT}

    ws  = wb.create_sheet("Feuil2")
    ws.sheet_properties.tabColor = _TAB_TCD
    cur = 1

    for soc in sorted(df[col_soc].astype(str).str.strip().unique()):
        sub   = df[df[col_soc].astype(str).str.strip() == soc][[col_motif, col_nb]]
        sub   = sub.sort_values(col_motif).reset_index(drop=True)
        total = sub[col_nb].sum()

        _write_section_filter(ws, cur, "Société", soc)
        cur += 2
        _write_header(ws, cur, HDRS)
        cur += 1
        for _, row in sub.iterrows():
            nb = row[col_nb]
            _write_row(ws, cur, [row[col_motif], nb, _pct(nb, total)], num_formats=NF)
            cur += 1
        _write_row(ws, cur, ["Total général", total, 1.0], total=True, num_formats=NF)
        cur += 2

    # Résumé global — toutes sociétés
    total_all = df[col_nb].sum()
    by_motif  = (
        df.groupby(col_motif, observed=True)[col_nb].sum().reset_index()
        .sort_values(col_motif)
    )
    _write_section_filter(ws, cur, "Société", "(Toutes)")
    cur += 2
    _write_header(ws, cur, HDRS)
    cur += 1
    for _, row in by_motif.iterrows():
        nb = row[col_nb]
        for col_idx, (val, fmt) in enumerate(
            [(row[col_motif], None), (nb, FMT_INT), (_pct(nb, total_all), FMT_PCT)], start=1
        ):
            cell = ws.cell(row=cur, column=col_idx, value=val)
            cell.font = _GRAND_TOTAL_FONT; cell.fill = _GRAND_TOTAL_FILL
            cell.alignment = _LEFT if col_idx == 1 else _CENTER
            if fmt:
                cell.number_format = fmt
        cur += 1
    for col_idx, (val, fmt) in enumerate(
        [("Total général", None), (total_all, FMT_INT), (1.0, FMT_PCT)], start=1
    ):
        cell = ws.cell(row=cur, column=col_idx, value=val)
        cell.font = _GRAND_TOTAL_FONT; cell.fill = _GRAND_TOTAL_FILL
        cell.alignment = _LEFT if col_idx == 1 else _CENTER
        if fmt:
            cell.number_format = fmt

    _auto_width(ws)
    socs = sorted(df[col_soc].astype(str).str.strip().unique())
    print(f"  [FEUIL2] {len(socs)} sociétés ({', '.join(socs)}) | résumé global {len(by_motif)} motifs")


def write_feuil3(wb, df: pd.DataFrame) -> None:
    """Table plate offre × codsoc × type_assure × nb_pp."""
    ws = wb.create_sheet("Feuil3")
    ws.sheet_properties.tabColor = _TAB_FLAT
    _write_header(ws, 1, list(df.columns))
    num_offsets = {i: FMT_INT for i, c in enumerate(df.columns)
                   if c.lower() in NUM_COLS_TYPE_AGE}
    for i, row in df.iterrows():
        _write_row(ws, i + 2, list(row), num_formats=num_offsets)
    ws.freeze_panes = "A2"
    _auto_width(ws)
    print(f"  [FEUIL3] {len(df)} lignes écrites")


def write_feuil4(wb, df: pd.DataFrame) -> None:
    """
    Feuil4 — deux blocs côte à côte :
      Gauche (cols A-E) : TCD codsoc × type_assure, une section par société.
      Droit  (cols F-K) : listing alphabétique de toutes les offres ASSURE
                          (aucune exclusion) avec total de regroupement en
                          formule (col I), libellé (col J) et ordre (col K).
    """
    col_soc   = next(c for c in df.columns if "codsoc" in c.lower())
    col_type  = next(c for c in df.columns if "type" in c.lower())
    col_date  = next(c for c in df.columns if "a date" in c.lower())
    col_nonr  = next(c for c in df.columns if "non radie" in c.lower())
    col_offre = next(c for c in df.columns if "offre" in c.lower())

    df = df.copy()
    df[col_soc] = _normalize_codsoc(df[col_soc])

    # Exclure les codsoc non-standard (ex: '000') du pivot par société
    df_valid = df[df[col_soc].astype(str).str.match(VALID_CODSOC)].copy()

    pivot_left = (
        df_valid.groupby([col_soc, col_type], dropna=False)[[col_date, col_nonr]]
        .sum(min_count=1).reset_index()
    )

    ws = wb.create_sheet("Feuil4")
    ws.sheet_properties.tabColor = _TAB_TCD
    HDRS_LEFT = ["Étiquettes de lignes", "Somme de nb pp a date",
                 "Somme de nb pp non radie", "% du total", "Taux radiation"]
    NF_LEFT = {1: FMT_INT, 2: FMT_INT}

    # ── Bloc gauche : codsoc × type_assure ───────────────────────────────────
    cur = 1
    for soc in sorted(pivot_left[col_soc].dropna().unique()):
        sub   = pivot_left[pivot_left[col_soc] == soc].sort_values(col_type)
        tot_d  = sub[col_date].sum()
        tot_nr = sub[col_nonr].sum()
        _write_section_filter(ws, cur, "codsoc", soc)
        cur += 2
        _write_header(ws, cur, HDRS_LEFT)
        cur += 1
        for _, row in sub.iterrows():
            pct  = _pct(row[col_nonr], tot_nr)
            trad = _taux_rad(row[col_date], row[col_nonr])
            _write_row(ws, cur, [row[col_type], row[col_date], row[col_nonr]], num_formats=NF_LEFT)
            for off, (val, fmt) in enumerate([(pct, FMT_PCT), (trad, FMT_PCT)], start=4):
                c = ws.cell(row=cur, column=off, value=val)
                c.font = _PCT_FONT; c.alignment = _CENTER; c.number_format = fmt
            cur += 1
        _write_row(ws, cur, ["Total général", tot_d, tot_nr, 1.0, _taux_rad(tot_d, tot_nr)],
                   total=True, num_formats={1: FMT_INT, 2: FMT_INT, 3: FMT_PCT, 4: FMT_PCT})
        cur += 2

    # Résumé global bloc gauche — toutes sociétés
    grand_d  = pivot_left[col_date].sum()
    grand_nr = pivot_left[col_nonr].sum()
    grand_by_type = (
        pivot_left.groupby(col_type)[[col_date, col_nonr]]
        .sum().reset_index().sort_values(col_type)
    )
    _write_section_filter(ws, cur, "codsoc", "(Toutes)")
    cur += 2
    _write_header(ws, cur, HDRS_LEFT)
    cur += 1
    for _, row in grand_by_type.iterrows():
        pct  = _pct(row[col_nonr], grand_nr)
        trad = _taux_rad(row[col_date], row[col_nonr])
        for col_idx, (val, fmt) in enumerate([
            (row[col_type], None), (row[col_date], FMT_INT), (row[col_nonr], FMT_INT),
            (pct, FMT_PCT), (trad, FMT_PCT),
        ], start=1):
            cell = ws.cell(row=cur, column=col_idx, value=val)
            cell.font = _GRAND_TOTAL_FONT if col_idx <= 3 else _PCT_FONT
            cell.fill = _GRAND_TOTAL_FILL
            cell.alignment = _LEFT if col_idx == 1 else _CENTER
            if fmt:
                cell.number_format = fmt
        cur += 1
    for col_idx, (val, fmt) in enumerate([
        ("Total général", None), (grand_d, FMT_INT), (grand_nr, FMT_INT),
        (1.0, FMT_PCT), (_taux_rad(grand_d, grand_nr), FMT_PCT),
    ], start=1):
        cell = ws.cell(row=cur, column=col_idx, value=val)
        cell.font = _GRAND_TOTAL_FONT; cell.fill = _GRAND_TOTAL_FILL
        cell.alignment = _LEFT if col_idx == 1 else _CENTER
        if fmt:
            cell.number_format = fmt
    cur += 2

    # ── Bloc droit : offre × nb_pp_non_radie (ASSURE) ────────────────────────
    # Listing alphabétique de TOUTES les offres ASSURE (aucune exclusion) avec,
    # pour les regroupements, le total en formule Excel (col I = =G..+G..),
    # le libellé du regroupement (col J) et l'ordre d'affichage (col K).
    RIGHT     = 6
    G_LETTER  = get_column_letter(RIGHT + 1)   # colonne "Somme de nb pp non radie"
    df_assure = df[df[col_type].astype(str).str.upper() == "ASSURE"]
    pivot_right = (
        df_assure.groupby(col_offre, dropna=False)[[col_nonr]]
        .sum(min_count=1).reset_index()
    )
    total_right = pivot_right[col_nonr].sum()

    # Regroupements recalculés sur les offres présentes : si une offre clé a
    # disparu du flux, son groupe est porté par le 1er membre encore présent.
    groupes = _groupes_effectifs({str(o) for o in pivot_right[col_offre].dropna()})

    # Tri des lignes de données par ordre d'affichage croissant (colonne K).
    # Les offres sans ordre (clé K vide) sont reléguées en bas du bloc.
    # NB : on trie l'ordre d'ÉCRITURE (et non les cellules a posteriori) afin que
    # les formules de regroupement (col I) référencent les bonnes lignes finales.
    rows_sorted = sorted(pivot_right.to_dict("records"),
                         key=lambda r: _tri_offre_key(r[col_offre], groupes))

    _write_section_filter(ws, 2, "type assure", "ASSURE", col_start=RIGHT)
    _write_header(ws, 4, ["Étiquettes de lignes", "Somme de nb pp non radie", "% du total ASSURE"],
                  col_start=RIGHT)
    # En-têtes des colonnes annexes (formule + libellé du regroupement + ordre)
    for col_off, title in ((3, "Formule à appliquer"),
                           (4, "Libellé du regroupement"),
                           (5, "Ordre d'affichage")):
        hcell = ws.cell(row=4, column=RIGHT + col_off, value=title)
        hcell.font = _HEADER_FONT; hcell.fill = _HEADER_FILL; hcell.alignment = _CENTER

    # 1ʳᵉ passe : écrire offre + somme + % (dans l'ordre trié) et mémoriser la
    # ligne finale de chaque offre.
    offre_rows: Dict[str, int] = {}
    for idx, row in enumerate(rows_sorted):
        offre    = row[col_offre]
        data_row = idx + 5
        offre_rows[offre] = data_row
        _write_row(ws, data_row, [offre, row[col_nonr]], col_start=RIGHT,
                   num_formats={1: FMT_INT})
        c = ws.cell(row=data_row, column=RIGHT + 2, value=_pct(row[col_nonr], total_right))
        c.font = _PCT_FONT; c.alignment = _CENTER; c.number_format = FMT_PCT

    # 2ᵉ passe : totaux de regroupement (colonne I = formule Excel =G..+G..),
    # libellé du regroupement (colonne J) et ordre d'affichage (colonne K).
    for offre, data_row in offre_rows.items():
        grp_num, grp_label = groupes.get(offre, (None, None))
        if grp_label:
            membres = [m.strip() for m in grp_label.split("+")]
            if len(membres) > 1:
                refs = [f"{G_LETTER}{offre_rows[m]}" for m in membres if m in offre_rows]
                if refs:
                    cell = ws.cell(row=data_row, column=RIGHT + 3, value="=" + "+".join(refs))
                    cell.font = _DATA_FONT; cell.number_format = FMT_INT
            ws.cell(row=data_row, column=RIGHT + 4, value=grp_label).font = _DATA_FONT
        if grp_num is not None:
            ws.cell(row=data_row, column=RIGHT + 5, value=grp_num).font = _DATA_FONT

    last = len(rows_sorted) + 5
    _write_row(ws, last, ["Total général", total_right, 1.0], total=True, col_start=RIGHT,
               num_formats={1: FMT_INT, 2: FMT_PCT})

    _auto_width(ws)
    print(f"  [FEUIL4] {len(pivot_left)} combinaisons codsoc×type | "
          f"{len(pivot_right)} offres ASSURE, total={int(total_right):,}")


def write_feuil7(wb, df: pd.DataFrame) -> None:
    """
    Feuil7 — listing à plat de TOUTES les offres ASSURE (aucune exclusion).
    Colonnes :
      A Étiquettes de lignes (offre)        B Somme de nb pp non radie
      C % du total ASSURE                   D Total (formule Excel : total du
        regroupement =B..+B.. porté par la 1ʳᵉ offre du groupe ; =B<ligne>
        pour une offre isolée disposant d'un ordre)
      E Formule (libellé du regroupement)   F Colonne (ordre d'affichage)
    Les offres membres d'un regroupement (hors clé) n'ont ni Total, ni Formule,
    ni ordre. Le total de regroupement les inclut via la formule de la clé.
    """
    col_type  = next(c for c in df.columns if "type" in c.lower())
    col_nonr  = next(c for c in df.columns if "non radie" in c.lower())
    col_offre = next(c for c in df.columns if "offre" in c.lower())

    df_assure = df[df[col_type].astype(str).str.upper() == "ASSURE"]
    pivot = (
        df_assure.groupby(col_offre, dropna=False)[[col_nonr]]
        .sum(min_count=1).reset_index()
    )
    total = pivot[col_nonr].sum()

    # Regroupements recalculés sur les offres présentes : si une offre clé a
    # disparu du flux, son groupe est porté par le 1er membre encore présent.
    groupes = _groupes_effectifs({str(o) for o in pivot[col_offre].dropna()})

    # Tri des lignes de données par ordre d'affichage croissant (colonne F) ;
    # offres sans ordre (clé F vide) reléguées en bas. Tri de l'ordre d'écriture
    # pour que les formules de la colonne D référencent les bonnes lignes.
    rows_sorted = sorted(pivot.to_dict("records"),
                         key=lambda r: _tri_offre_key(r[col_offre], groupes))

    # Offres "membres" d'un regroupement (toutes sauf la porteuse) : pas de
    # ligne de synthèse propre (Total / Formule / ordre).
    membres_groupes = set()
    for key, (_grp_num, grp_label) in groupes.items():
        if grp_label:
            for m in (x.strip() for x in grp_label.split("+")):
                if m != key:
                    membres_groupes.add(m)

    ws = wb.create_sheet("Feuil7")
    ws.sheet_properties.tabColor = _TAB_FLAT
    _write_header(ws, 1, ["Étiquettes de lignes", "Somme de nb pp non radie",
                          "% du total ASSURE", "Total", "Formule", "Colonne"])

    offre_rows = {r[col_offre]: idx + 2 for idx, r in enumerate(rows_sorted)}

    for idx, row in enumerate(rows_sorted):
        offre = row[col_offre]
        r     = idx + 2
        nonr  = row[col_nonr]
        _write_row(ws, r, [offre, nonr], num_formats={1: FMT_INT})
        c = ws.cell(row=r, column=3, value=_pct(nonr, total))
        c.font = _PCT_FONT; c.alignment = _CENTER; c.number_format = FMT_PCT

        if offre in membres_groupes:
            continue  # offre membre : agrégée dans le total de sa clé

        grp_num, grp_label = groupes.get(offre, (None, None))
        membres = [m.strip() for m in grp_label.split("+")] if grp_label else []
        dval = None
        if len(membres) > 1:
            refs = [f"B{offre_rows[m]}" for m in membres if m in offre_rows]
            if refs:
                dval = "=" + "+".join(refs)
            ws.cell(row=r, column=5, value=grp_label).font = _DATA_FONT
        elif grp_num is not None:
            dval = f"=B{r}"
        if dval is not None:
            dcell = ws.cell(row=r, column=4, value=dval)
            dcell.font = _DATA_FONT; dcell.alignment = _CENTER; dcell.number_format = FMT_INT
        if grp_num is not None:
            oc = ws.cell(row=r, column=6, value=grp_num)
            oc.font = _DATA_FONT; oc.alignment = _CENTER

    # Ligne de totaux (ordre = max + 1, cf. modèle)
    tr = len(rows_sorted) + 2
    _write_row(ws, tr, ["Total général", total, 1.0], total=True,
               num_formats={1: FMT_INT, 2: FMT_PCT})
    dtot = ws.cell(row=tr, column=4, value=f"=B{tr}")
    dtot.font = _TOTAL_FONT; dtot.fill = _TOTAL_FILL
    dtot.alignment = _CENTER; dtot.number_format = FMT_INT
    # Ordre du total = max des ordres réellement affichés + 1 (les offres
    # configurées mais absentes du flux ne créent pas de trou de numérotation).
    orders = [g[0] for g in groupes.values() if g[0] is not None]
    if orders:
        ftot = ws.cell(row=tr, column=6, value=max(orders) + 1)
        ftot.font = _TOTAL_FONT; ftot.fill = _TOTAL_FILL; ftot.alignment = _CENTER

    ws.freeze_panes = "A2"
    _auto_width(ws)
    print(f"  [FEUIL7] {len(pivot)} offres ASSURE (aucune exclusion), total={int(total):,}")


def write_feuil5(wb, df: pd.DataFrame) -> None:
    """Table plate offre × codsoc × tranche_âge × nb_pp."""
    ws = wb.create_sheet("Feuil5")
    ws.sheet_properties.tabColor = _TAB_FLAT
    _write_header(ws, 1, list(df.columns))
    num_offsets = {i: FMT_INT for i, c in enumerate(df.columns)
                   if c.lower() in NUM_COLS_TYPE_AGE}
    for i, row in df.iterrows():
        _write_row(ws, i + 2, list(row), num_formats=num_offsets)
    ws.freeze_panes = "A2"
    _auto_width(ws)
    print(f"  [FEUIL5] {len(df)} lignes écrites")


def write_feuil6(wb, df: pd.DataFrame) -> None:
    """TCD : tranche_âge par codsoc — sections par société + KPIs + résumé."""
    col_soc  = next(c for c in df.columns if "codsoc" in c.lower())
    col_age  = next(c for c in df.columns if "tranche" in c.lower())
    col_date = next(c for c in df.columns if "a date" in c.lower())
    col_nonr = next(c for c in df.columns if "non radie" in c.lower())

    df = df.copy()
    df[col_soc] = _normalize_codsoc(df[col_soc])

    pivot = (
        df.groupby([col_soc, col_age], dropna=False)[[col_date, col_nonr]]
        .sum(min_count=1).reset_index()
    )

    HDRS = ["Étiquettes de lignes", "Somme de nb pp a date",
            "Somme de nb pp non radie", "% du total société", "Taux radiation"]
    NF   = {1: FMT_INT, 2: FMT_INT}

    ws = wb.create_sheet("Feuil6")
    ws.sheet_properties.tabColor = _TAB_TCD
    cur = 1

    for soc in sorted(pivot[col_soc].dropna().unique()):
        sub    = pivot[pivot[col_soc] == soc].sort_values(col_age)
        tot_d  = sub[col_date].sum()
        tot_nr = sub[col_nonr].sum()

        _write_section_filter(ws, cur, "codsoc", soc)
        cur += 2
        _write_header(ws, cur, HDRS)
        cur += 1
        for _, row in sub.iterrows():
            pct  = _pct(row[col_nonr], tot_nr)
            trad = _taux_rad(row[col_date], row[col_nonr])
            _write_row(ws, cur, [row[col_age], row[col_date], row[col_nonr]], num_formats=NF)
            for off, (val, fmt) in enumerate([(pct, FMT_PCT), (trad, FMT_PCT)], start=4):
                c = ws.cell(row=cur, column=off, value=val)
                c.font = _PCT_FONT; c.alignment = _CENTER; c.number_format = fmt
            cur += 1
        _write_row(ws, cur, ["Total général", tot_d, tot_nr, 1.0, _taux_rad(tot_d, tot_nr)],
                   total=True, num_formats={1: FMT_INT, 2: FMT_INT, 3: FMT_PCT, 4: FMT_PCT})
        cur += 2

    # Résumé global — toutes sociétés
    grand_d  = pivot[col_date].sum()
    grand_nr = pivot[col_nonr].sum()
    grand_by_age = (
        pivot.groupby(col_age)[[col_date, col_nonr]]
        .sum().reset_index().sort_values(col_age)
    )
    _write_section_filter(ws, cur, "codsoc", "(Toutes)")
    cur += 2
    _write_header(ws, cur, HDRS)
    cur += 1
    for _, row in grand_by_age.iterrows():
        pct  = _pct(row[col_nonr], grand_nr)
        trad = _taux_rad(row[col_date], row[col_nonr])
        for col_idx, (val, fmt) in enumerate([
            (row[col_age], None), (row[col_date], FMT_INT), (row[col_nonr], FMT_INT),
            (pct, FMT_PCT), (trad, FMT_PCT),
        ], start=1):
            cell = ws.cell(row=cur, column=col_idx, value=val)
            cell.font = _GRAND_TOTAL_FONT if col_idx <= 3 else _PCT_FONT
            cell.fill = _GRAND_TOTAL_FILL
            cell.alignment = _LEFT if col_idx == 1 else _CENTER
            if fmt:
                cell.number_format = fmt
        cur += 1
    for col_idx, (val, fmt) in enumerate([
        ("Total général", None), (grand_d, FMT_INT), (grand_nr, FMT_INT),
        (1.0, FMT_PCT), (_taux_rad(grand_d, grand_nr), FMT_PCT),
    ], start=1):
        cell = ws.cell(row=cur, column=col_idx, value=val)
        cell.font = _GRAND_TOTAL_FONT; cell.fill = _GRAND_TOTAL_FILL
        cell.alignment = _LEFT if col_idx == 1 else _CENTER
        if fmt:
            cell.number_format = fmt

    _auto_width(ws)
    print(f"  [FEUIL6] {len(pivot)} combinaisons, {pivot[col_soc].nunique()} sociétés + résumé global")


def write_synthese(wb, df_type: pd.DataFrame, df_resil: pd.DataFrame,
                   df_age: pd.DataFrame, prefix: str,
                   kpi_scalars: Optional[Dict[str, float]] = None,
                   kpi8: Optional[Dict[str, float]] = None) -> None:
    """
    Feuille Synthèse : indicateurs clés issus des 3 blocs de données
    + KPI scalaires et déphasages extraits de la feuille source.
    Sections : KPI source | Portefeuille | Déphasages KPI 8 | Résiliations
               | Top offres | Tranche fidèle | Par type assuré | Par société
    """
    # Colonnes sources
    col_soc_t  = next(c for c in df_type.columns if "codsoc" in c.lower())
    col_type   = next(c for c in df_type.columns if "type" in c.lower())
    col_date_t = next(c for c in df_type.columns if "a date" in c.lower())
    col_nonr_t = next(c for c in df_type.columns if "non radie" in c.lower())
    col_offre  = next(c for c in df_type.columns if "offre" in c.lower())

    col_motif  = next(c for c in df_resil.columns if "motif" in c.lower())
    col_nb_r   = next(c for c in df_resil.columns if "nombre" in c.lower())

    col_soc_a  = next(c for c in df_age.columns if "codsoc" in c.lower())
    col_age    = next(c for c in df_age.columns if "tranche" in c.lower())
    col_date_a = next(c for c in df_age.columns if "a date" in c.lower())
    col_nonr_a = next(c for c in df_age.columns if "non radie" in c.lower())

    # Préparation
    dt = df_type.copy()
    dt[col_soc_t] = _normalize_codsoc(dt[col_soc_t])
    # Filtrer les codsoc non-standard pour être cohérent avec l'onglet Feuil4
    dt = dt[dt[col_soc_t].astype(str).str.match(VALID_CODSOC)].copy()

    dr = df_resil.copy()
    dr[col_nb_r]  = pd.to_numeric(dr[col_nb_r], errors="coerce")
    dr[col_motif] = dr[col_motif].fillna("").astype(str).str.strip()
    # Les résiliations sans motif renseigné sont regroupées sous "(vide)" pour
    # être incluses dans le Total général (cohérence avec Feuil2 et TCD source).
    dr.loc[dr[col_motif] == "", col_motif] = RESIL_EMPTY_LABEL
    # Filtrer aux codes DWH + "(vide)" (cohérence avec Feuil2)
    dr = dr[dr[col_motif].isin(RESILIATION_ORDER_WITH_EMPTY)].copy()

    da = df_age.copy()
    da[col_soc_a] = _normalize_codsoc(da[col_soc_a])

    # ── KPI globaux ──────────────────────────────────────────────────────────
    total_nonr  = dt[col_nonr_t].sum()
    total_date  = dt[col_date_t].sum()
    taux_rad_g  = _taux_rad(total_date, total_nonr)
    total_resil = dr[col_nb_r].sum()

    # Nombre de contrats ASSURE non radiés (cohérent avec Feuil4)
    total_assure_nonr = dt.loc[
        dt[col_type].astype(str).str.upper() == "ASSURE", col_nonr_t
    ].sum()

    # Motif dominant (toutes sociétés) — calculé hors "(vide)" : les résiliations
    # sans motif renseigné ne peuvent pas être qualifiées de "motif dominant".
    motif_grp_all = dr.groupby(col_motif, observed=True)[col_nb_r].sum()
    motif_grp     = motif_grp_all.drop(RESIL_EMPTY_LABEL, errors="ignore")
    motif_top     = motif_grp.idxmax()
    motif_top_n   = motif_grp.max()

    # Top 3 offres ASSURE par pp non radie
    df_assure  = dt[dt[col_type].astype(str).str.upper() == "ASSURE"]
    top_offres = (
        df_assure.groupby(col_offre)[col_nonr_t].sum()
        .sort_values(ascending=False).head(3)
    )

    # Tranche d'âge la plus fidèle (taux radiation le plus bas)
    age_pivot = da.groupby(col_age)[[col_date_a, col_nonr_a]].sum().reset_index()
    age_pivot["_trad"] = age_pivot.apply(
        lambda r: _taux_rad(r[col_date_a], r[col_nonr_a]) if r[col_nonr_a] else 1.0,
        axis=1,
    )
    tranche_fidele = age_pivot.loc[age_pivot["_trad"].idxmin(), col_age]
    tranche_taux   = age_pivot["_trad"].min()

    # Par type assuré
    type_pivot = (
        dt.groupby(col_type)[[col_date_t, col_nonr_t]].sum().reset_index()
        .sort_values(col_nonr_t, ascending=False)
    )

    # Par société
    soc_pivot = (
        dt.groupby(col_soc_t)[[col_date_t, col_nonr_t]].sum().reset_index()
        .sort_values(col_soc_t)
    )

    # ── Écriture ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Synthèse")
    ws.sheet_properties.tabColor = _TAB_SYNTH

    _TITLE_FONT   = Font(bold=True, size=14, color="1F497D")
    _SECTION_FONT = Font(bold=True, size=11, color="1F497D")
    _SECT_FILL    = PatternFill("solid", fgColor="DCE6F1")
    _LABEL_FONT   = Font(size=10)
    _VAL_FONT     = Font(bold=True, size=10)

    def _section(r: int, title: str) -> int:
        cell = ws.cell(row=r, column=1, value=title)
        cell.font = _SECTION_FONT
        cell.fill = _SECT_FILL
        return r + 1

    def _kv(r: int, label: str, value, fmt: Optional[str] = None) -> int:
        ws.cell(row=r, column=1, value=label).font = _LABEL_FONT
        c = ws.cell(row=r, column=2, value=value)
        c.font = _VAL_FONT; c.alignment = _CENTER
        if fmt:
            c.number_format = fmt
        return r + 1

    cur = 1

    # Titre
    ws.cell(row=cur, column=1,
            value=f"Synthèse KPI Accolade — flux {prefix}").font = _TITLE_FONT
    cur += 2

    # KPI issus de la feuille source (si disponibles)
    if kpi_scalars:
        _KPI_LABELS = {
            # Libellés alignés sur le vocabulaire du flux source depuis la V1.14
            # (17/07/2026) : "AI" y est devenu "Accolade".
            "kpi1_nouvelles_souscriptions":  ("KPI 1 — Nouvelles souscriptions Accolade depuis 01/01/2024", FMT_INT),
            "kpi2_resiliations_total":        ("KPI 2 — Résiliations Accolade toutes dates",                FMT_INT),
            "kpi2bis_resiliations_2024":      ("KPI 2 Bis — Résiliations Accolade depuis 01/01/2024",       FMT_INT),
            "kpi5_sans_kpep_2024":            ("KPI 5 — Assurés sans KPEP rapproché (depuis 2024)",         FMT_INT),
            "kpi5bis_sans_kpep_total":        ("KPI 5 Bis — Assurés sans KPEP rapproché (total)",           FMT_INT),
            "kpi6_sans_mail":                 ("KPI 6 — Assurés principaux sans mail dans Accolade",        FMT_INT),
        }
        cur = _section(cur, "KPI issus de la feuille source")
        for key, (label, fmt) in _KPI_LABELS.items():
            if key in kpi_scalars:
                cur = _kv(cur, label, kpi_scalars[key], fmt)
        cur += 1

    # Portefeuille
    cur = _section(cur, "Portefeuille")
    cur = _kv(cur, "Total pp non radiés",     total_nonr,  FMT_INT)
    cur = _kv(cur, "Total pp à date",          total_date,  FMT_INT)
    cur = _kv(cur, "Taux de radiation global", taux_rad_g,  FMT_PCT)
    cur += 1

    # Qualité des données — Déphasages KPI 8
    if kpi8:
        _SUBTOTAL_FILL = PatternFill("solid", fgColor="EBF1DE")   # vert pâle — sous-total tél
        _SUBTOTAL_FONT = Font(bold=True, italic=True, size=10, color="375623")

        cur = _section(cur, "Qualité des données — Déphasages AI / KPEP (KPI 8)")
        _write_header(ws, cur, ["Indicateur", "Nb déphasages", "% portefeuille ASSURE"],
                      col_start=1)
        # Forcer largeur col 1 plus large pour les labels
        cur += 1

        _DEP_ROWS = [
            ("dep_mail",       "Email"),
            ("dep_tel_port",   "Téléphone PORT"),
            ("dep_tel_telp",   "Téléphone TELP"),
            ("dep_naissance",  "Date de naissance"),
            ("dep_deces",      "Date de décès"),
            ("dep_adresses",   "Adresses"),
        ]
        for key, label in _DEP_ROWS:
            nb = kpi8.get(key)
            if nb is None:
                continue
            pct_val = _pct(nb, total_assure_nonr)
            ws.cell(row=cur, column=1, value=label).font = _DATA_FONT
            c_nb  = ws.cell(row=cur, column=2, value=nb)
            c_pct = ws.cell(row=cur, column=3, value=pct_val)
            c_nb.font  = _DATA_FONT;  c_nb.alignment  = _CENTER; c_nb.number_format  = FMT_INT
            c_pct.font = _PCT_FONT;   c_pct.alignment = _CENTER; c_pct.number_format = FMT_PCT
            # Insérer la ligne sous-total téléphone après TELP
            if key == "dep_tel_telp":
                cur += 1
                tel_total = kpi8.get("dep_tel_port", 0) + kpi8.get("dep_tel_telp", 0)
                pct_tel   = _pct(tel_total, total_assure_nonr)
                for col_idx, (val, fmt) in enumerate(
                    [("  ↳ Total téléphone", None), (tel_total, FMT_INT), (pct_tel, FMT_PCT)],
                    start=1,
                ):
                    cell = ws.cell(row=cur, column=col_idx, value=val)
                    cell.font = _SUBTOTAL_FONT
                    cell.fill = _SUBTOTAL_FILL
                    cell.alignment = _LEFT if col_idx == 1 else _CENTER
                    if fmt:
                        cell.number_format = fmt
            cur += 1
        cur += 1

    # Résiliations
    cur = _section(cur, "Résiliations")
    cur = _kv(cur, "Total résiliations",     total_resil,                     FMT_INT)
    cur = _kv(cur, "Motif dominant",          motif_top)
    cur = _kv(cur, "Part du motif dominant",  _pct(motif_top_n, total_resil), FMT_PCT)
    cur += 1

    # Top 3 offres ASSURE
    cur = _section(cur, "Top 3 offres ASSURE (pp non radiés)")
    for offre, val in top_offres.items():
        cur = _kv(cur, str(offre), val, FMT_INT)
    cur += 1

    # Tranche d'âge la plus fidèle
    cur = _section(cur, "Fidélité par tranche d'âge")
    cur = _kv(cur, "Tranche la plus fidèle", tranche_fidele)
    cur = _kv(cur, "Taux radiation",          tranche_taux,  FMT_PCT)
    cur += 1

    # Par type assuré
    cur = _section(cur, "Par type assuré")
    _write_header(ws, cur, ["Type assuré", "pp non radiés", "% total", "Taux radiation"])
    cur += 1
    for _, row in type_pivot.iterrows():
        pct  = _pct(row[col_nonr_t], total_nonr)
        trad = _taux_rad(row[col_date_t], row[col_nonr_t])
        for col_idx, (val, fmt) in enumerate([
            (row[col_type], None), (row[col_nonr_t], FMT_INT),
            (pct, FMT_PCT), (trad, FMT_PCT),
        ], start=1):
            cell = ws.cell(row=cur, column=col_idx, value=val)
            cell.font = _DATA_FONT
            cell.alignment = _LEFT if col_idx == 1 else _CENTER
            if fmt:
                cell.number_format = fmt
        cur += 1
    cur += 1

    # Par société
    cur = _section(cur, "Par société")
    _write_header(ws, cur, ["Société", "pp non radiés", "% total", "Taux radiation"])
    cur += 1
    for _, row in soc_pivot.iterrows():
        pct  = _pct(row[col_nonr_t], total_nonr)
        trad = _taux_rad(row[col_date_t], row[col_nonr_t])
        for col_idx, (val, fmt) in enumerate([
            (row[col_soc_t], None), (row[col_nonr_t], FMT_INT),
            (pct, FMT_PCT), (trad, FMT_PCT),
        ], start=1):
            cell = ws.cell(row=cur, column=col_idx, value=val)
            cell.font = _DATA_FONT
            cell.alignment = _LEFT if col_idx == 1 else _CENTER
            if fmt:
                cell.number_format = fmt
        cur += 1

    _auto_width(ws)
    trad_str = f"{taux_rad_g:.2%}" if taux_rad_g is not None else "—"
    top_n    = _pct(motif_top_n, total_resil)
    top_str  = f"{top_n:.2%}" if top_n is not None else "—"
    print(f"  [SYNTHESE] Taux rad global={trad_str} | "
          f"Total résil={int(total_resil):,} | "
          f"Motif dominant='{motif_top}' ({top_str})")

# ─── KPI 4 — DOUBLONS DE CONTRATS ─────────────────────────────────────────────

# Patterns dans la zone d'en-tête (rows 1..17) → clé dict.
# Les valeurs sont sur la même ligne ou la ligne précédente (cf. fichier source
# où la valeur "Prévoyance" apparaît juste au-dessus de son libellé).
_KPI4_HEADER_PATTERNS: List[tuple] = [
    ("plusieurs contrats (santé,prévoyance)", "kpi4_total"),
    ("plusieurs contrats prévoyance",          "kpi4_prevoyance"),
    ("plusieurs contrats santé",                "kpi4_sante"),
]


def find_kpi4_input_file() -> Optional[Path]:
    """Recherche le fichier 'Accolade - KPI 4*.xlsx' dans INPUT_DIR.

    Le fichier KPI principal ('Accolade - KPI*', hors 'KPI 4') est exclu :
    il ne doit jamais être confondu avec un fichier KPI 4, sous peine d'écraser
    l'export TCD déjà généré (même nom de sortie) puis d'échouer au parsing.
    """
    matches = [
        f for f in INPUT_DIR.iterdir()
        if KPI4_FILENAME_RE.match(f.name) and not KPI_FILENAME_RE.match(f.name)
    ]
    if not matches:
        return None
    if len(matches) > 1:
        matches.sort(key=lambda f: KPI4_FILENAME_RE.match(f.name).group(1), reverse=True)
        names = [f.name for f in matches]
        warnings.warn(
            f"[WARN] Plusieurs fichiers KPI 4 détectés : {names}\n"
            f"  → Utilisation du plus récent : {matches[0].name}",
            stacklevel=2,
        )
    return matches[0]


def extract_prefix_kpi4(filename: str) -> str:
    m = KPI4_FILENAME_RE.match(filename)
    if not m:
        raise ValueError(f"Impossible d'extraire le préfixe depuis '{filename}'")
    return m.group(1)


def find_kpi4_main_sheet(wb) -> str:
    """Retourne le nom de la feuille principale du fichier KPI 4 (ex: 'Accolade - KPI 4 V2.0')."""
    for name in wb.sheetnames:
        if KPI4_MAIN_SHEET_RE.match(name):
            return name
    raise ValueError(
        f"Aucune feuille 'Accolade - KPI 4*' trouvée. Feuilles disponibles : {wb.sheetnames}"
    )


def parse_kpi4_header(ws) -> Dict[str, float]:
    """
    Extrait les 3 KPI scalaires depuis la zone d'en-tête (rows 1..17).
    La valeur cible est cherchée sur la même ligne, puis sur les lignes
    voisines (±2) si absente — le fichier source place parfois la valeur
    une ligne au-dessus du libellé.
    """
    result: Dict[str, float] = {}
    title_rows: Dict[str, int] = {}

    # 1ʳᵉ passe : localiser les libellés
    for row_idx in range(1, KPI4_HEADER_SCAN_ROWS + 1):
        for col_idx in range(1, ws.max_column + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if not v:
                continue
            text = str(v).strip().lower()
            for pattern, key in _KPI4_HEADER_PATTERNS:
                if key not in title_rows and pattern in text:
                    title_rows[key] = row_idx

    # 2ᵉ passe : valeur numérique sur la ligne du titre, sinon ±2 lignes
    for key, row_idx in title_rows.items():
        for delta in (0, -1, 1, -2, 2):
            r = row_idx + delta
            if r < 1 or r > KPI4_HEADER_SCAN_ROWS:
                continue
            for col_idx in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=col_idx).value
                if isinstance(v, (int, float)) and v > 0 and not isinstance(v, bool):
                    result[key] = float(v)
                    break
            if key in result:
                break
    return result


def parse_kpi4_data(ws) -> pd.DataFrame:
    """
    Extrait le bloc de données (en-tête à KPI4_DATA_HEADER_ROW) avec :
      nom long, prenom, code insee, cle insee, date naissance,
      code société, type prestation, count.
    """
    col_indices: List[int] = []
    headers: List[str] = []
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(row=KPI4_DATA_HEADER_ROW, column=col_idx).value
        if v is not None:
            col_indices.append(col_idx)
            headers.append(str(v).strip())

    if not headers:
        raise ValueError(
            f"En-tête introuvable à la ligne {KPI4_DATA_HEADER_ROW} de '{ws.title}'"
        )

    rows = []
    for row_idx in range(KPI4_DATA_HEADER_ROW + 1, ws.max_row + 1):
        row_data = [ws.cell(row=row_idx, column=c).value for c in col_indices]
        if all(v is None for v in row_data):
            continue
        rows.append(row_data)

    df = pd.DataFrame(rows, columns=headers).dropna(how="all").reset_index(drop=True)
    if "count" in df.columns:
        df["count"] = pd.to_numeric(df["count"], errors="coerce")
    return df


def write_kpi_doublon_contrat(wb, df: pd.DataFrame, prefix: str,
                              kpi_header: Dict[str, float]) -> None:
    """
    Génère la feuille 'KPI_doublon_contrat' avec :
      - KPI sources (header)
      - Synthèse globale
      - Répartition par type prestation
      - Répartition par code société
      - Croisement société × type prestation (TCD)
      - Distribution du nombre de contrats par assuré
      - Top 10 cas extrêmes
      - Assurés multi-types
    """
    # Localisation des colonnes
    col_nom    = next((c for c in df.columns if "nom" in c.lower()), None)
    col_prenom = next((c for c in df.columns if "prenom" in c.lower()), None)
    col_insee  = next((c for c in df.columns if "code insee" in c.lower()), None)
    col_cle    = next((c for c in df.columns if "cle" in c.lower()), None)
    col_naiss  = next((c for c in df.columns if "naissance" in c.lower()), None)
    col_soc    = next((c for c in df.columns if "soci" in c.lower()), None)
    col_type   = next((c for c in df.columns if "type" in c.lower()), None)
    col_count  = next((c for c in df.columns if c.lower() == "count"), None)

    missing = [n for n, c in [
        ("nom long", col_nom), ("prenom", col_prenom), ("code insee", col_insee),
        ("date naissance", col_naiss), ("code société", col_soc),
        ("type prestation", col_type), ("count", col_count),
    ] if c is None]
    if missing:
        raise ValueError(
            f"[KPI 4] Colonnes manquantes dans le bloc de données : {missing}\n"
            f"  Colonnes détectées : {list(df.columns)}"
        )

    df = df.copy()
    df[col_count] = pd.to_numeric(df[col_count], errors="coerce").fillna(0).astype(int)
    df[col_soc]   = df[col_soc].astype(str).str.strip().str.zfill(3)
    df[col_type]  = df[col_type].astype(str).str.strip().str.upper()

    # Métriques globales
    nb_assures      = len(df)
    nb_contrats_tot = int(df[col_count].sum())
    moy_contrats    = (nb_contrats_tot / nb_assures) if nb_assures else None

    # Pivot type prestation
    by_type = (
        df.groupby(col_type)
          .agg(nb_assures=(col_count, "size"), nb_contrats=(col_count, "sum"))
          .reset_index().sort_values("nb_assures", ascending=False)
    )

    # Pivot code société
    by_soc = (
        df.groupby(col_soc)
          .agg(nb_assures=(col_count, "size"), nb_contrats=(col_count, "sum"))
          .reset_index().sort_values(col_soc)
    )

    # Croisement société × type prestation (matrice de comptages d'assurés)
    cross = (
        df.pivot_table(index=col_soc, columns=col_type, values=col_count,
                       aggfunc="size", fill_value=0)
          .reset_index().sort_values(col_soc)
    )
    type_cols = [c for c in cross.columns if c != col_soc]

    # Distribution du nombre de contrats par assuré
    dist = (
        df.groupby(col_count)
          .size().reset_index(name="nb_assures")
          .sort_values(col_count)
    )
    dist["nb_contrats"] = dist[col_count] * dist["nb_assures"]

    # Top 10 cas extrêmes (count décroissant)
    top_cols = [c for c in (col_nom, col_prenom, col_naiss, col_soc, col_type, col_count)
                if c is not None]
    top10 = df.sort_values(col_count, ascending=False).head(10)[top_cols].reset_index(drop=True)

    # Assurés multi-types : même personne (nom, prenom, insee, naissance) avec
    # plusieurs types de prestation distincts dans le fichier
    pers_keys = [c for c in (col_nom, col_prenom, col_insee, col_cle, col_naiss) if c is not None]
    multi = (
        df.groupby(pers_keys)[col_type].nunique().reset_index(name="nb_types_distincts")
    )
    multi = multi[multi["nb_types_distincts"] >= 2]
    nb_multi_types = len(multi)

    # ── Écriture de la feuille ───────────────────────────────────────────────
    if KPI4_OUTPUT_SHEET in wb.sheetnames:
        del wb[KPI4_OUTPUT_SHEET]
    ws = wb.create_sheet(KPI4_OUTPUT_SHEET)
    ws.sheet_properties.tabColor = _TAB_SYNTH

    _TITLE_FONT   = Font(bold=True, size=14, color="1F497D")
    _SECTION_FONT = Font(bold=True, size=11, color="1F497D")
    _SECT_FILL    = PatternFill("solid", fgColor="DCE6F1")
    _LABEL_FONT   = Font(size=10)
    _VAL_FONT     = Font(bold=True, size=10)

    def _section(r: int, title: str) -> int:
        cell = ws.cell(row=r, column=1, value=title)
        cell.font = _SECTION_FONT
        cell.fill = _SECT_FILL
        return r + 1

    def _kv(r: int, label: str, value, fmt: Optional[str] = None) -> int:
        ws.cell(row=r, column=1, value=label).font = _LABEL_FONT
        c = ws.cell(row=r, column=2, value=value)
        c.font = _VAL_FONT; c.alignment = _CENTER
        if fmt:
            c.number_format = fmt
        return r + 1

    cur = 1
    ws.cell(row=cur, column=1,
            value=f"KPI 4 — Doublons de contrats — flux {prefix}").font = _TITLE_FONT
    cur += 2

    # KPI sources (header du fichier)
    if kpi_header:
        cur = _section(cur, "KPI issus du fichier source")
        labels = {
            "kpi4_total":       "KPI 4 — Assurés avec plusieurs contrats (Santé + Prévoyance)",
            "kpi4_prevoyance":  "KPI 4 — Assurés avec plusieurs contrats Prévoyance",
            "kpi4_sante":       "KPI 4 — Assurés avec plusieurs contrats Santé",
        }
        for key, label in labels.items():
            if key in kpi_header:
                cur = _kv(cur, label, kpi_header[key], FMT_INT)
        cur += 1

    # Synthèse globale
    cur = _section(cur, "Synthèse globale (issue du bloc de données)")
    cur = _kv(cur, "Nombre d'assurés principaux", nb_assures, FMT_INT)
    cur = _kv(cur, "Nombre total de contrats en doublon", nb_contrats_tot, FMT_INT)
    cur = _kv(cur, "Nombre moyen de contrats par assuré", moy_contrats, "0.00")
    cur = _kv(cur, "Assurés ayant plusieurs types de prestations distincts",
              nb_multi_types, FMT_INT)
    cur += 1

    # Répartition par type prestation
    cur = _section(cur, "Répartition par type de prestation")
    _write_header(ws, cur,
                  ["Type prestation", "Nb assurés", "% assurés",
                   "Nb contrats", "% contrats"])
    cur += 1
    for _, row in by_type.iterrows():
        pct_a = _pct(row["nb_assures"],  nb_assures)
        pct_c = _pct(row["nb_contrats"], nb_contrats_tot)
        for col_idx, (val, fmt) in enumerate([
            (row[col_type], None),
            (int(row["nb_assures"]),  FMT_INT),
            (pct_a, FMT_PCT),
            (int(row["nb_contrats"]), FMT_INT),
            (pct_c, FMT_PCT),
        ], start=1):
            cell = ws.cell(row=cur, column=col_idx, value=val)
            cell.font = _DATA_FONT
            cell.alignment = _LEFT if col_idx == 1 else _CENTER
            if fmt:
                cell.number_format = fmt
        cur += 1
    _write_row(ws, cur, ["Total général", nb_assures, 1.0, nb_contrats_tot, 1.0],
               total=True, num_formats={1: FMT_INT, 2: FMT_PCT, 3: FMT_INT, 4: FMT_PCT})
    cur += 2

    # Répartition par code société
    cur = _section(cur, "Répartition par code société")
    _write_header(ws, cur,
                  ["Code société", "Nb assurés", "% assurés",
                   "Nb contrats", "% contrats"])
    cur += 1
    for _, row in by_soc.iterrows():
        pct_a = _pct(row["nb_assures"],  nb_assures)
        pct_c = _pct(row["nb_contrats"], nb_contrats_tot)
        for col_idx, (val, fmt) in enumerate([
            (row[col_soc], None),
            (int(row["nb_assures"]),  FMT_INT),
            (pct_a, FMT_PCT),
            (int(row["nb_contrats"]), FMT_INT),
            (pct_c, FMT_PCT),
        ], start=1):
            cell = ws.cell(row=cur, column=col_idx, value=val)
            cell.font = _DATA_FONT
            cell.alignment = _LEFT if col_idx == 1 else _CENTER
            if fmt:
                cell.number_format = fmt
        cur += 1
    _write_row(ws, cur, ["Total général", nb_assures, 1.0, nb_contrats_tot, 1.0],
               total=True, num_formats={1: FMT_INT, 2: FMT_PCT, 3: FMT_INT, 4: FMT_PCT})
    cur += 2

    # Croisement société × type prestation
    cur = _section(cur, "Croisement société × type de prestation (nb assurés)")
    _write_header(ws, cur, ["Code société", *type_cols, "Total"])
    cur += 1
    col_totals = {tc: 0 for tc in type_cols}
    for _, row in cross.iterrows():
        line_total = int(sum(row[tc] for tc in type_cols))
        vals = [row[col_soc]] + [int(row[tc]) for tc in type_cols] + [line_total]
        nf = {i: FMT_INT for i in range(1, len(vals))}
        _write_row(ws, cur, vals, num_formats=nf)
        for tc in type_cols:
            col_totals[tc] += int(row[tc])
        cur += 1
    grand_total_cross = sum(col_totals.values())
    total_vals = ["Total général"] + [col_totals[tc] for tc in type_cols] + [grand_total_cross]
    _write_row(ws, cur, total_vals, total=True,
               num_formats={i: FMT_INT for i in range(1, len(total_vals))})
    cur += 2

    # Distribution du nombre de contrats par assuré
    cur = _section(cur, "Distribution du nombre de contrats par assuré")
    _write_header(ws, cur,
                  ["Nb contrats", "Nb assurés", "% assurés",
                   "Total contrats", "% contrats"])
    cur += 1
    for _, row in dist.iterrows():
        pct_a = _pct(row["nb_assures"],  nb_assures)
        pct_c = _pct(row["nb_contrats"], nb_contrats_tot)
        for col_idx, (val, fmt) in enumerate([
            (int(row[col_count]),     FMT_INT),
            (int(row["nb_assures"]),  FMT_INT),
            (pct_a, FMT_PCT),
            (int(row["nb_contrats"]), FMT_INT),
            (pct_c, FMT_PCT),
        ], start=1):
            cell = ws.cell(row=cur, column=col_idx, value=val)
            cell.font = _DATA_FONT
            cell.alignment = _CENTER
            if fmt:
                cell.number_format = fmt
        cur += 1
    _write_row(ws, cur, ["Total général", nb_assures, 1.0, nb_contrats_tot, 1.0],
               total=True, num_formats={1: FMT_INT, 2: FMT_PCT, 3: FMT_INT, 4: FMT_PCT})
    cur += 2

    # Top 10 cas extrêmes
    cur = _section(cur, "Top 10 — assurés ayant le plus de contrats")
    _write_header(ws, cur, top_cols)
    cur += 1
    for _, row in top10.iterrows():
        vals = [row[c] for c in top_cols]
        nf   = {len(top_cols) - 1: FMT_INT}   # dernier col = count
        _write_row(ws, cur, vals, num_formats=nf)
        cur += 1
    cur += 1

    # Assurés multi-types (échantillon)
    cur = _section(cur, f"Assurés ayant plusieurs types de prestations distincts ({nb_multi_types})")
    if nb_multi_types > 0:
        sample_cols = pers_keys + ["nb_types_distincts"]
        _write_header(ws, cur, sample_cols)
        cur += 1
        for _, row in multi.head(20).iterrows():
            vals = [row[c] for c in sample_cols]
            nf   = {len(sample_cols) - 1: FMT_INT}
            _write_row(ws, cur, vals, num_formats=nf)
            cur += 1

    ws.freeze_panes = "A2"
    _auto_width(ws)
    print(f"  [KPI_doublon_contrat] {nb_assures} assurés, {nb_contrats_tot} contrats, "
          f"{nb_multi_types} multi-types")


def process_kpi4_file() -> None:
    """Traite le fichier 'Accolade - KPI 4*.xlsx' s'il est présent dans Input_Data."""
    input_file = find_kpi4_input_file()
    if input_file is None:
        print(f"\n[INFO]    Aucun fichier '*_Accolade - KPI 4*.xlsx' trouvé dans {INPUT_DIR}.")
        print(f"          → Étape KPI 4 ignorée.")
        return

    prefix      = extract_prefix_kpi4(input_file.name)
    output_file = OUTPUT_DIR / input_file.name

    print(f"\n[KPI 4]   {input_file.name}  (préfixe: {prefix})")
    print(f"[OUTPUT]  {output_file}")

    # Contrôle préalable AVANT toute copie : le fichier doit réellement contenir
    # une feuille principale 'Accolade - KPI 4*'. Sinon on saute proprement —
    # sans écraser un éventuel export existant ni interrompre le pipeline.
    # NB : openpyxl.Workbook ne supporte pas le protocole de gestionnaire de
    # contexte (pas de `with`) ; on garantit la fermeture via try/finally.
    try:
        wb_check = load_workbook(input_file, read_only=True)
        try:
            sheet_names = list(wb_check.sheetnames)
        finally:
            wb_check.close()
    except Exception as exc:
        print(f"[WARN]    Impossible d'ouvrir '{input_file.name}' ({exc}).")
        print(f"          → Étape KPI 4 ignorée.")
        return
    if not any(KPI4_MAIN_SHEET_RE.match(n) for n in sheet_names):
        print(f"[WARN]    Aucune feuille 'Accolade - KPI 4*' dans '{input_file.name}' "
              f"(feuilles : {sheet_names}).")
        print(f"          → Étape KPI 4 ignorée.")
        return

    wb = None
    try:
        shutil.copy2(input_file, output_file)
        print(f"[COPY]    Fichier dupliqué vers Output/")

        wb = load_workbook(output_file)
        main_sheet_name = find_kpi4_main_sheet(wb)
        ws_main = wb[main_sheet_name]
        print(f"[LOAD]    Feuille '{main_sheet_name}' chargée "
              f"({ws_main.max_row} lignes × {ws_main.max_column} colonnes)")

        kpi_header = parse_kpi4_header(ws_main)
        print(f"  [HEADER] {len(kpi_header)} KPI extraits : "
              + ", ".join(f"{k}={int(v):,}" for k, v in kpi_header.items()))

        df = parse_kpi4_data(ws_main)
        print(f"  [DATA]   {len(df)} lignes extraites, colonnes : {list(df.columns)}")

        write_kpi_doublon_contrat(wb, df, prefix, kpi_header)

        wb.save(output_file)
        print(f"[DONE]    {output_file.name} sauvegardé avec la feuille '{KPI4_OUTPUT_SHEET}'.")
    except Exception as exc:
        print(f"[WARN]    Échec du traitement KPI 4 de '{input_file.name}' : {exc}")
        print(f"          → Étape KPI 4 ignorée (le reste du pipeline n'est pas affecté).")
    finally:
        # openpyxl ne supporte pas `with` : fermeture explicite garantie même
        # en cas d'exception (libère le verrou fichier, notamment sous Windows).
        if wb is not None:
            wb.close()


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main() -> None:
    print("=" * 60)
    print("05_generation_tcd.py — Génération TCD Accolade KPI")
    print("=" * 60)

    # 1. Localisation du fichier source (facultatif — l'absence n'est pas une erreur)
    input_file = find_input_file()
    if input_file is None:
        print(f"\n[INFO]    Aucun fichier '*_Accolade - KPI*.xlsx' trouvé dans {INPUT_DIR}.")
        print(f"          → Étape TCD V1.12 ignorée (skip uniquement si fichier absent).")
        # On poursuit pour traiter éventuellement le fichier KPI 4
        process_kpi4_file()
        print("=" * 60)
        return

    prefix      = extract_prefix(input_file.name)
    output_file = OUTPUT_DIR / input_file.name

    print(f"\n[INPUT]   {input_file.name}  (préfixe: {prefix})")
    print(f"[OUTPUT]  {output_file}")

    # 2. Copie vers Output/
    shutil.copy2(input_file, output_file)
    print("\n[COPY]    Fichier dupliqué vers Output/")

    # 3. Chargement du workbook
    wb = load_workbook(output_file)
    main_sheet = find_main_sheet(wb)
    ws_main = wb[main_sheet]
    print(f"[LOAD]    Feuille '{main_sheet}' chargée "
          f"({ws_main.max_row} lignes × {ws_main.max_column} colonnes)")

    # 4. Parsing des blocs de données
    print("\n[PARSER]  Extraction des blocs...")

    print(f"  Recherche '{MARKER_RESILIATIONS}'...", end=" ")
    row_resil = find_marker_row(ws_main, MARKER_RESILIATIONS)
    print(f"ligne {row_resil}")
    df_resil = extract_block(ws_main, row_resil, NUM_COLS_RESIL)
    validate_block(df_resil, "résiliations")
    print(f"  [OK] {len(df_resil)} lignes, schéma validé")

    print(f"  Recherche '{MARKER_TYPE_ASSURE}'...", end=" ")
    row_type = find_marker_row(ws_main, MARKER_TYPE_ASSURE)
    print(f"ligne {row_type}")
    df_type = extract_block(ws_main, row_type, NUM_COLS_TYPE_AGE)
    validate_block(df_type, "type_assure")
    print(f"  [OK] {len(df_type)} lignes, schéma validé")

    print(f"  Recherche '{MARKER_TRANCHE_AGE}'...", end=" ")
    row_age = find_marker_row(ws_main, MARKER_TRANCHE_AGE)
    print(f"ligne {row_age}")
    df_age = extract_block(ws_main, row_age, NUM_COLS_TYPE_AGE)
    validate_block(df_age, "tranche_âge")
    print(f"  [OK] {len(df_age)} lignes, schéma validé")

    # 4b. Extraction des KPI scalaires et déphasages de la feuille source
    print("\n[CHECK]   Contrôles de cohérence...")
    kpi_scalars = parse_kpi_scalars(ws_main)
    print(f"  [KPI SOURCE] {len(kpi_scalars)} indicateurs extraits : "
          f"{list(kpi_scalars.keys())}")
    kpi_manquants = sorted(_KPI_SCALAR_REQUIRED - set(kpi_scalars))
    if kpi_manquants:
        print(
            f"\n  [WARN] KPI source introuvables : {kpi_manquants}\n"
            f"    Les intitulés de la feuille '{ws_main.title}' ont probablement changé.\n"
            f"    Le bloc 'KPI issus de la feuille source' de la Synthèse sera incomplet\n"
            f"    (voire absent). → Vérifier _KPI_SCALAR_PATTERNS."
        )
    kpi8 = parse_kpi8_dephasages(ws_main)
    if kpi8:
        print(f"  [KPI 8]  {len(kpi8)} déphasages extraits : "
              + ", ".join(f"{k}={int(v):,}" for k, v in kpi8.items()))

    # 4c. Validation de robustesse (offres, cross-check total)
    validate_data(df_type)

    # Cross-check total pp non radie ASSURE vs KPI 7 (référence source)
    col_type_chk = next(c for c in df_type.columns if "type" in c.lower())
    col_nonr_chk = next(c for c in df_type.columns if "non radie" in c.lower())
    total_calc = df_type[df_type[col_type_chk].astype(str).str.upper() == "ASSURE"][col_nonr_chk].sum()
    # KPI 7 est le total "par offre" : on peut l'approcher via kpi1 si présent
    # La valeur de référence exacte est dans Feuil4 col F total (calculé)
    print(f"  [CHECK] Total ASSURE calculé = {int(total_calc):,}  "
          f"(référence manuelle KPI 7 attendue ≈ 2 338 292)")

    # 5. Suppression des anciennes feuilles TCD
    print("\n[CLEAN]   Suppression des anciens TCD...")
    for sheet_name in TCD_SHEETS:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            print(f"  Feuille '{sheet_name}' supprimée")

    # 6. Génération des nouvelles feuilles
    print("\n[TCD]     Génération des feuilles...")
    write_synthese(wb, df_type, df_resil, df_age, prefix,
                   kpi_scalars=kpi_scalars, kpi8=kpi8)
    write_feuil1(wb, df_resil)
    write_feuil2(wb, df_resil)
    write_feuil3(wb, df_type)
    write_feuil4(wb, df_type)
    write_feuil7(wb, df_type)
    write_feuil5(wb, df_age)
    write_feuil6(wb, df_age)

    # 7. Sauvegarde
    wb.save(output_file)
    print(f"\n[DONE]    {output_file.name} sauvegardé avec succès.")

    # 8. Traitement complémentaire — fichier KPI 4 (doublons de contrats)
    print("\n" + "─" * 60)
    print("[KPI 4]   Recherche du fichier 'Accolade - KPI 4*.xlsx'")
    print("─" * 60)
    process_kpi4_file()
    print("=" * 60)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"\n[ERREUR]  {exc}", file=sys.stderr)
        sys.exit(1)